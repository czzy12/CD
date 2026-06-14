import re
from html.parser import HTMLParser
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import Transaction
from .number_parser import balance_candidates, extract_signed_amount, money_to_decimal


BANK_NAME = "Excel导入"
CENT = Decimal("0.01")


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[tuple[Any, ...]]] = []
        self._current_table: list[tuple[Any, ...]] | None = None
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._current_table = []
        elif tag == "tr" and self._current_table is not None:
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            self._current_row.append("".join(self._current_cell).strip())
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None and self._current_table is not None:
            self._current_table.append(tuple(self._current_row))
            self._current_row = None
        elif tag == "table" and self._current_table is not None:
            self.tables.append(self._current_table)
            self._current_table = None


def _norm(value: Any) -> str:
    return str(value or "").replace("\n", "").replace(" ", "").replace("　", "").strip()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def _read_html_excel_tables(path: str) -> list[list[tuple[Any, ...]]]:
    raw = Path(path).read_bytes()
    text = raw.decode("utf-8", errors="ignore")
    parser = _HtmlTableParser()
    parser.feed(text)
    return parser.tables


def _find_col(headers: list[str], names: tuple[str, ...], exclude: set[int] | None = None) -> int | None:
    exclude = exclude or set()
    for name in names:
        for index, header in enumerate(headers):
            if index not in exclude and header == name:
                return index
    for name in names:
        for index, header in enumerate(headers):
            if index not in exclude and name in header:
                return index
    return None


def _parse_date_part(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_text(value).replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _parse_time_part(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    text = _cell_text(value).replace("：", ":").replace("；", ":").replace(";", ":")
    match = re.search(r"(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?", text)
    if not match:
        return time(0, 0, 0)
    try:
        return time(int(match.group(1)), int(match.group(2)), int(match.group(3) or "0"))
    except ValueError:
        return time(0, 0, 0)


def _normalize_ocr_date_text(value: Any) -> str:
    text = _cell_text(value).strip()
    replacements = str.maketrans({
        "G": "6",
        "g": "6",
        "E": "6",
        "e": "6",
        "D": "0",
        "d": "0",
        "O": "0",
        "o": "0",
        "I": "1",
        "l": "1",
    })
    return text.translate(replacements)


def _parse_datetime(date_value: Any, time_value: Any | None = None) -> datetime | None:
    if isinstance(date_value, datetime) and time_value in (None, ""):
        return date_value.replace(microsecond=0)

    if time_value not in (None, ""):
        parsed_date = _parse_date_part(_normalize_ocr_date_text(date_value))
        if parsed_date is None:
            return None
        return datetime.combine(parsed_date, _parse_time_part(time_value))

    text = _normalize_ocr_date_text(date_value).replace("：", ":").replace("/", "-")
    text = text.replace("年", "-").replace("月", "-").replace("日", " ")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?", text)
    if not match:
        return None
    try:
        parsed_date = date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        parsed_time = time(int(match.group(4) or "0"), int(match.group(5) or "0"), int(match.group(6) or "0"))
        return datetime.combine(parsed_date, parsed_time)
    except ValueError:
        return None


def _parse_money(value: Any) -> Decimal | None:
    if isinstance(value, Decimal):
        return value.quantize(CENT)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return Decimal(str(value)).quantize(CENT)
        except InvalidOperation:
            return None
    text = _cell_text(value)
    if not text:
        return None
    cleaned = (
        text.replace("￥", "")
        .replace("元", "")
        .replace("，", ",")
        .replace("。", ".")
        .replace("．", ".")
        .replace(" ", "")
    )
    direct = money_to_decimal(cleaned)
    if direct is not None:
        return direct
    signed = extract_signed_amount(cleaned)
    if signed is not None:
        return signed
    try:
        return Decimal(re.sub(r"[^0-9.\-]", "", cleaned)).quantize(CENT)
    except (InvalidOperation, ValueError):
        return None


def _direction(text: str) -> str | None:
    compact = _norm(text)
    if any(word in compact for word in ("收入", "贷", "转入", "入账", "收")):
        return "income"
    if any(word in compact for word in ("支出", "借", "转出", "出账", "付")):
        return "expense"
    return None


def _resolve_missing_directions(transactions: list[Transaction]) -> None:
    previous: Transaction | None = None
    for tx in sorted(transactions, key=lambda item: (item.transaction_time, item.row_no)):
        if previous is not None and previous.balance is not None and tx.balance is not None:
            amount = _parse_money(tx.raw_amount)
            if amount is not None and tx.income == 0 and tx.expense == 0:
                if (previous.balance + amount).quantize(CENT) == tx.balance.quantize(CENT):
                    tx.income = amount
                    tx.issues = [issue for issue in tx.issues if issue != "收支方向无法解析"]
                elif (previous.balance - amount).quantize(CENT) == tx.balance.quantize(CENT):
                    tx.expense = amount
                    tx.issues = [issue for issue in tx.issues if issue != "收支方向无法解析"]
                if not tx.issues:
                    tx.status = "ok"
        previous = tx


def _header_mapping(rows: list[tuple[Any, ...]]) -> tuple[int, list[str], dict[str, int]] | None:
    for row_index, row in enumerate(rows[:30]):
        headers = [_norm(cell) for cell in row]
        time_col = _find_col(headers, ("交易时间", "交易日期时间", "日期时间"))
        date_col = time_col if time_col is not None else _find_col(headers, ("交易日期", "记账日期", "日期"))
        separate_time_col = None if time_col is not None else _find_col(headers, ("时间", "交易时刻"))
        amount_col = _find_col(headers, ("交易金额", "发生额", "本次金额", "交易额", "金额"))
        income_col = _find_col(headers, ("收入金额", "收入", "贷方发生额", "贷方", "贷"))
        expense_col = _find_col(headers, ("支出金额", "支出", "借方发生额", "借方", "借"))
        direction_col = _find_col(headers, ("收入/支出", "收入支出", "收支", "借贷标志", "借贷状态", "借贷方向", "方向"))

        exclude = {col for col in (amount_col, income_col, expense_col) if col is not None}
        balance_col = _find_col(headers, ("账户余额", "本次余额", "交易余额", "余额", "金额"), exclude)

        if date_col is not None and (amount_col is not None or income_col is not None or expense_col is not None):
            return row_index, headers, {
                "date": date_col,
                "time": separate_time_col,
                "amount": amount_col,
                "income": income_col,
                "expense": expense_col,
                "direction": direction_col,
                "balance": balance_col,
            }
    return None


def _is_boc_converted_sheet(rows: list[tuple[Any, ...]]) -> bool:
    sample = "\n".join(" ".join(_cell_text(cell) for cell in row[:8]) for row in rows[:8])
    if "中国银行交易流水明细清单" in sample and "记账日期" in sample and "金额" in sample and "余额" in sample:
        return True
    for row in rows[:12]:
        headers = [_norm(cell) for cell in row]
        if all(header in headers for header in ("记账日期", "记账时间", "币别", "金额", "余额", "交易名称")):
            return True
    return False


def _extract_boc_converted_sheet(rows: list[tuple[Any, ...]], sheet_index: int) -> list[Transaction]:
    header_mapping = None
    for row_index, row in enumerate(rows[:30]):
        headers = [_norm(cell) for cell in row]
        date_col = _find_col(headers, ("记账日期",))
        time_col = _find_col(headers, ("记账时间",))
        currency_col = _find_col(headers, ("币别",))
        amount_col = _find_col(headers, ("金额",))
        balance_col = _find_col(headers, ("余额",))
        if None not in (date_col, time_col, currency_col, amount_col, balance_col):
            header_mapping = (row_index, headers, date_col, time_col, currency_col, amount_col, balance_col)
            break
    if header_mapping is None:
        return []

    header_row, headers, date_col, time_col, currency_col, amount_col, balance_col = header_mapping
    transactions: list[Transaction] = []
    for excel_row_index, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        currency = _norm(row[currency_col] if currency_col < len(row) else "")
        if currency != "人民币":
            continue

        tx_time = _parse_datetime(
            row[date_col] if date_col < len(row) else None,
            row[time_col] if time_col < len(row) else None,
        )
        amount = _parse_money(row[amount_col] if amount_col < len(row) else None)
        balance = _parse_money(row[balance_col] if balance_col < len(row) else None)
        if tx_time is None or amount is None or balance is None:
            continue

        income = amount if amount >= 0 else Decimal("0.00")
        expense = -amount if amount < 0 else Decimal("0.00")
        raw_fields = [_cell_text(cell) for cell in row]
        tx = Transaction(
            # BOC converted sheets are printed newest-first. When two rows have
            # the same second, row order must be reversed to restore the chain.
            transaction_time=tx_time.replace(microsecond=max(0, 999999 - excel_row_index)),
            income=income.quantize(CENT),
            expense=expense.quantize(CENT),
            balance=balance.quantize(CENT),
            bank="中国银行Excel导入",
            page_no=sheet_index,
            row_no=excel_row_index,
            raw_time=f"{_cell_text(row[date_col])} {_cell_text(row[time_col])}",
            raw_amount=_cell_text(row[amount_col]),
            raw_balance=_cell_text(row[balance_col]),
            raw_text=" | ".join(raw_fields),
            raw_fields=raw_fields,
            raw_headers=headers,
        )
        tx.preserve_signed_columns = True
        transactions.append(tx)

    _resolve_boc_converted_balances(transactions)
    return transactions


def _balance_cents(value: Decimal) -> str:
    return f"{value.quantize(CENT):.2f}".replace(".", "")


def _raw_balance_digits(raw: str) -> str:
    normalized = _normalize_ocr_date_text(raw)
    return re.sub(r"\D", "", normalized)


def _boc_balance_candidates(raw: str) -> list[Decimal]:
    candidates: list[Decimal] = []
    seen: set[Decimal] = set()

    def add(value: Decimal | None) -> None:
        if value is not None and value >= Decimal("0.00") and value not in seen:
            candidates.append(value.quantize(CENT))
            seen.add(value.quantize(CENT))

    normalized = _normalize_ocr_date_text(raw)
    add(_parse_money(normalized))
    for candidate in balance_candidates(normalized):
        add(candidate)

    digits = _raw_balance_digits(raw)
    if "." not in normalized and len(digits) > 4:
        try:
            add((Decimal(digits) / Decimal("100")).quantize(CENT))
        except InvalidOperation:
            pass
    return candidates


def _expected_supported_by_raw(expected: Decimal, raw: str, candidates: list[Decimal]) -> bool:
    expected_digits = _balance_cents(expected)
    raw_digits = _raw_balance_digits(raw)
    if expected in candidates:
        return True
    if any(expected_digits.endswith(_balance_cents(candidate)) for candidate in candidates):
        return True
    # PDF-to-Excel OCR can drop one digit from the balance. Only repair when
    # nearly all raw digits appear in the expected balance in order.
    if raw_digits and len(raw_digits) >= 5:
        position = 0
        for char in expected_digits:
            if position < len(raw_digits) and raw_digits[position] == char:
                position += 1
        if position >= len(raw_digits) - 1:
            return True
    return False


def _resolve_boc_converted_balances(transactions: list[Transaction]) -> None:
    previous_balance: Decimal | None = None
    for tx in sorted(transactions, key=lambda item: (item.transaction_time, item.page_no, item.row_no)):
        amount = (tx.income - tx.expense).quantize(CENT)
        candidates = _boc_balance_candidates(tx.raw_balance)
        if previous_balance is not None:
            expected = (previous_balance + amount).quantize(CENT)
            if _expected_supported_by_raw(expected, tx.raw_balance, candidates):
                tx.balance = expected
        elif candidates:
            tx.balance = candidates[0]
        if tx.balance is not None:
            previous_balance = tx.balance.quantize(CENT)


def _extract_sheet(rows: list[tuple[Any, ...]], sheet_index: int) -> list[Transaction]:
    if _is_boc_converted_sheet(rows):
        return _extract_boc_converted_sheet(rows, sheet_index)

    mapping = _header_mapping(rows)
    if mapping is None:
        return []

    header_row, headers, cols = mapping
    transactions: list[Transaction] = []
    for excel_row_index, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        tx_time = _parse_datetime(row[cols["date"]] if cols["date"] is not None and cols["date"] < len(row) else None,
                                  row[cols["time"]] if cols["time"] is not None and cols["time"] < len(row) else None)
        if tx_time is None:
            continue

        raw_amount = _cell_text(row[cols["amount"]]) if cols["amount"] is not None and cols["amount"] < len(row) else ""
        raw_balance = _cell_text(row[cols["balance"]]) if cols["balance"] is not None and cols["balance"] < len(row) else ""
        raw_income = _cell_text(row[cols["income"]]) if cols["income"] is not None and cols["income"] < len(row) else ""
        raw_expense = _cell_text(row[cols["expense"]]) if cols["expense"] is not None and cols["expense"] < len(row) else ""
        if not any((raw_amount, raw_income, raw_expense, raw_balance)):
            continue

        amount = _parse_money(raw_amount)
        balance = _parse_money(raw_balance)
        issues: list[str] = []

        income = _parse_money(row[cols["income"]]) if cols["income"] is not None and cols["income"] < len(row) else None
        expense = _parse_money(row[cols["expense"]]) if cols["expense"] is not None and cols["expense"] < len(row) else None
        if income is not None or expense is not None:
            income = income or Decimal("0.00")
            expense = expense or Decimal("0.00")
            raw_amount = raw_amount or _cell_text(income if income > 0 else expense)
        elif amount is not None:
            raw_direction = _cell_text(row[cols["direction"]]) if cols["direction"] is not None and cols["direction"] < len(row) else ""
            direction = _direction(raw_direction)
            if direction == "income":
                income = amount
                expense = Decimal("0.00")
            elif direction == "expense":
                income = Decimal("0.00")
                expense = amount
            elif amount < 0:
                income = Decimal("0.00")
                expense = abs(amount)
            else:
                income = Decimal("0.00")
                expense = Decimal("0.00")
                issues.append("收支方向无法解析")
        else:
            income = Decimal("0.00")
            expense = Decimal("0.00")
            issues.append("交易金额无法解析")

        if balance is None and cols["balance"] is not None:
            issues.append("余额无法解析")

        tx = Transaction(
            transaction_time=tx_time,
            income=income,
            expense=expense,
            balance=balance,
            bank=BANK_NAME,
            page_no=sheet_index,
            row_no=excel_row_index,
            raw_time=_cell_text(row[cols["date"]]),
            raw_amount=raw_amount,
            raw_balance=raw_balance,
            raw_text=" | ".join(_cell_text(cell) for cell in row),
            raw_fields=[_cell_text(cell) for cell in row],
            raw_headers=headers,
            status="ok" if not issues else "review",
            issues=issues,
        )
        tx.preserve_signed_columns = True
        tx.balance_tolerance = Decimal("0.99")
        transactions.append(tx)

    _resolve_missing_directions(transactions)
    return transactions


def extract_excel_transactions(path: str) -> list[Transaction]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException):
        best: list[Transaction] = []
        for sheet_index, rows in enumerate(_read_html_excel_tables(path), start=1):
            transactions = _extract_sheet(rows, sheet_index)
            if len(transactions) > len(best):
                best = transactions
        return best

    best: list[Transaction] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        rows = list(worksheet.iter_rows(values_only=True))
        transactions = _extract_sheet(rows, sheet_index)
        if len(transactions) > len(best):
            best = transactions
    return best
