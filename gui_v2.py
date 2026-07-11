import sys
import re
import os
import subprocess
from dataclasses import dataclass
from datetime import datetime, time
from decimal import Decimal
from pathlib import Path

from PyQt6.QtCore import QDate, QRectF, QThread, QTimer, Qt, pyqtSignal
from PyQt6.QtGui import QColor, QKeySequence, QPainter, QPainterPath, QPalette, QPen
from PyQt6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QCheckBox,
    QDateEdit,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QLineEdit,
    QScrollArea,
    QStatusBar,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from bankflow_v2.adjustment import AdjustmentConfig, AdjustmentResult, apply_adjustments, parse_amount_wan
from bankflow_v2.income_proof_export import (
    balance_wechat_summaries,
    flow_type as income_flow_type,
    looks_corporate_account_name,
    write_income_proof_input,
    write_salary_income_proof_input,
)
from bankflow_v2.summary import Issue, Summary, money, monthly_summaries, sort_transactions, summarize


SUPPORTED_INPUTS = {".pdf", ".xlsx", ".xlsm"}
DATE_RANGE_EMPTY_MESSAGE = "日期范围内没有流水"
FIXED_PROOF_MONTHS = 6
SALARY_KEYWORDS = ("工资", "代发", "薪资", "薪酬", "奖金")
ACCOUNT_NAME_PATTERNS = (
    r"户名\s*(?:Account Name)?\s*[:：]\s*([^\s，,]+)",
    r"客户姓名\s*[:：]\s*([^\s，,]+)",
    r"账户名称\s*[:：]\s*([^\s，,]+)",
    r"户主\s*[:：]\s*([^\s，,]+)",
    r"Account Name\s*[:：]\s*([^\s，,]+)",
)
ACCOUNT_NO_PATTERNS = (
    r"客户账号\s*(?:Account No\.?|Account Number)?\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"账号/卡号\s*(?:Account/Card No\.?)?\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"Account/Card\s*No\.?\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"账号\s*(?:Account No\.?|Account Number)?\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"账户\s*(?:Account No\.?|Account Number)?\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"卡号\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"银行卡号\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
    r"Account\s*(?:No\.?|Number)\s*[:：]?\s*([0-9][0-9\s\-*＊]{6,30}[0-9])",
)


def is_packaged_app() -> bool:
    return getattr(sys, "frozen", False) or "__compiled__" in globals()


def runtime_dir() -> Path:
    if is_packaged_app():
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def child_env() -> dict[str, str]:
    env = os.environ.copy()
    for key in (
        "PYTHONHOME",
        "PYTHONPATH",
        "QT_PLUGIN_PATH",
        "QT_QPA_PLATFORM_PLUGIN_PATH",
        "QTWEBENGINEPROCESS_PATH",
    ):
        env.pop(key, None)
    return env


@dataclass
class FileResult:
    path: Path
    bank_id: str
    bank_label: str
    bank_confidence: int
    bank_reason: str
    summary: object
    transactions: list
    status: str
    message: str
    account_name: str = ""
    account_no: str = ""


@dataclass
class ConfidenceInfo:
    score: int
    level: str
    tone: str
    reason: str


class DropTable(QTableWidget):
    filesDropped = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.viewport().setAcceptDrops(True)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)
        self.setShowGrid(False)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setStretchLastSection(False)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.horizontalScrollBar().setStyleSheet(
            """
            QScrollBar:horizontal {
                background: #ffffff;
                border: 0;
                height: 10px;
                margin: 0;
            }
            QScrollBar::handle:horizontal {
                background: #cfd7e3;
                border-radius: 5px;
                min-width: 42px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #9aa6b2;
            }
            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal,
            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: transparent;
                border: 0;
                width: 0;
            }
            """
        )

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        paths = [Path(url.toLocalFile()) for url in event.mimeData().urls()]
        self.filesDropped.emit(paths)

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.StandardKey.Copy):
            self.copy_selection()
            return
        super().keyPressEvent(event)

    def copy_selection(self):
        ranges = self.selectedRanges()
        if not ranges:
            return

        lines: list[str] = []
        for selected in ranges:
            for row in range(selected.topRow(), selected.bottomRow() + 1):
                values = []
                for col in range(selected.leftColumn(), selected.rightColumn() + 1):
                    item = self.item(row, col)
                    values.append(item.text() if item else "")
                lines.append("\t".join(values))
        QApplication.clipboard().setText("\n".join(lines))


class RoundedCornerOverlay(QWidget):
    def __init__(self, parent: QWidget) -> None:
        super().__init__(parent)
        self._outside_color = QColor("#ffffff")
        self._border_color = QColor("#dbe8fb")
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

    def set_colors(self, outside_color: str, border_color: str) -> None:
        self._outside_color = QColor(outside_color)
        self._border_color = QColor(border_color)
        self.update()

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        rounded_rect = QRectF(self.rect()).adjusted(0.5, 1.5, -0.5, -0.5)
        rounded_path = QPainterPath()
        rounded_path.addRoundedRect(rounded_rect, 12, 12)
        outside_path = QPainterPath()
        outside_path.addRect(QRectF(self.rect()))
        painter.fillPath(outside_path.subtracted(rounded_path), self._outside_color)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(QPen(self._border_color, 1))
        painter.drawPath(rounded_path)


class RoundedTableShell(QFrame):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._corner_overlay: RoundedCornerOverlay | None = None

    def set_corner_colors(self, outside_color: str, border_color: str) -> None:
        if self._corner_overlay is None:
            self._corner_overlay = RoundedCornerOverlay(self)
        self._corner_overlay.set_colors(outside_color, border_color)
        self._corner_overlay.setGeometry(self.rect())
        self._corner_overlay.show()
        self._corner_overlay.raise_()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        if self._corner_overlay is not None:
            self._corner_overlay.setGeometry(self.rect())
            self._corner_overlay.raise_()


class DropWidget(QWidget):
    filesDropped = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        paths = [Path(url.toLocalFile()) for url in event.mimeData().urls()]
        self.filesDropped.emit(paths)


class Worker(QThread):
    finished = pyqtSignal(list, list)
    progress = pyqtSignal(str)

    def __init__(
        self,
        paths: list[Path],
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        pdf_passwords: dict[Path, str] | None = None,
    ):
        super().__init__()
        self.paths = paths
        self.start_date = start_date
        self.end_date = end_date
        self.pdf_passwords = pdf_passwords or {}

    def _filter_transactions(self, transactions: list) -> list:
        if self.start_date is None and self.end_date is None:
            return transactions

        filtered = []
        for tx in transactions:
            if self.start_date is not None and tx.transaction_time < self.start_date:
                continue
            if self.end_date is not None and tx.transaction_time > self.end_date:
                continue
            filtered.append(tx)
        return filtered

    def _generic_pdf_label(self, detection) -> str:
        if detection.bank_id and detection.label not in ("未识别", ""):
            return f"{detection.label}（通用识别）"
        return "通用PDF识别"

    def _extract_with_fallback(self, path: Path, detection) -> tuple[list, str, str, bool]:
        if path.suffix.lower() in (".xlsx", ".xlsm"):
            from bankflow_v2.excel_input import extract_excel_transactions

            return extract_excel_transactions(str(path)), "Excel导入", "Excel文件导入", False

        if not detection.bank_id:
            from bankflow_v2.generic_pdf import extract_generic_pdf

            transactions = extract_generic_pdf(str(path))
            if transactions:
                return transactions, "通用PDF识别", f"{detection.reason}，已使用通用识别", True
            return [], "未识别", detection.reason, False

        if detection.bank_id in ("cmbc", "cib", "generic_pdf"):
            from bankflow_v2.pipeline import extract_transactions

            transactions = extract_transactions(str(path), detection.bank_id)
            return transactions, self._generic_pdf_label(detection), "已使用通用识别", True

        try:
            from bankflow_v2.pipeline import extract_transactions

            transactions = extract_transactions(str(path), detection.bank_id)
        except Exception as exc:
            from bankflow_v2.generic_pdf import extract_generic_pdf

            fallback = extract_generic_pdf(str(path))
            if fallback:
                return fallback, self._generic_pdf_label(detection), f"专用解析失败：{exc}；已使用通用识别", True
            raise

        if transactions:
            return transactions, detection.label, "", False

        from bankflow_v2.generic_pdf import extract_generic_pdf

        fallback = extract_generic_pdf(str(path))
        if fallback:
            return fallback, self._generic_pdf_label(detection), "专用解析未得到流水，已使用通用识别", True
        return transactions, detection.label, "未解析到流水", False

    def run(self):
        from bankflow_v2.pdf_password import install_pdf_password_support, register_pdf_passwords

        install_pdf_password_support()
        register_pdf_passwords(self.pdf_passwords)

        results: list[FileResult] = []
        all_issues: list[Issue] = []

        for path in self.paths:
            self.progress.emit(f"处理中: {path.name}")
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                detection = type("Detection", (), {
                    "bank_id": "excel",
                    "label": "Excel导入",
                    "confidence": 100,
                    "reason": "Excel文件导入",
                })()
            else:
                from bankflow_v2.auto_detect import detect_bank_type

                detection = detect_bank_type(str(path))

            try:
                transactions, bank_label, fallback_message, used_generic = self._extract_with_fallback(path, detection)
                account_name = extract_account_name(path)
                account_no = extract_account_no(path)
                bank_label = infer_excel_bank_label(bank_label, transactions)
                detected_flow_type = infer_flow_type(detection.bank_id, account_name, transactions)
                original_count = len(transactions)
                transactions = self._filter_transactions(transactions)
                for tx in transactions:
                    tx.source_file = path.name
                    tx.bank_label = bank_label
                    tx.flow_type = detected_flow_type
                file_summary = summarize(transactions, path.name)
                all_issues.extend(file_summary.issues)
                review_issues = [issue for issue in file_summary.issues if issue.level == "需复核"]
                if transactions and not review_issues:
                    status = "通用识别" if used_generic else "正常"
                else:
                    status = "需复核"
                message = fallback_message if transactions else (fallback_message or "未解析到流水")
                if original_count and not transactions:
                    message = DATE_RANGE_EMPTY_MESSAGE
                if message and (not transactions or review_issues):
                    all_issues.append(Issue("需复核", path.name, "", message))
                results.append(
                    FileResult(
                        path,
                        detection.bank_id,
                        bank_label,
                        detection.confidence,
                        detection.reason,
                        file_summary,
                        transactions,
                        status,
                        message,
                        account_name,
                        account_no,
                    )
                )
            except Exception as exc:
                issue = Issue("需复核", path.name, "", f"解析失败: {exc}")
                all_issues.append(issue)
                results.append(FileResult(path, detection.bank_id, detection.label, detection.confidence, detection.reason, summarize([]), [], "需复核", str(exc), extract_account_name(path), extract_account_no(path)))

        self.finished.emit(results, all_issues)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("银行流水 PDF/Excel 解析")
        self.resize(1120, 680)
        self.setMinimumSize(960, 600)
        self.paths: list[Path] = []
        self.results: list[FileResult] = []
        self.issues: list[Issue] = []
        self.worker: Worker | None = None
        self.adjustment_result = AdjustmentResult()
        self.cached_transactions = []
        self.setAcceptDrops(True)
        self.adjustment_refresh_timer = QTimer(self)
        self.adjustment_refresh_timer.setSingleShot(True)
        self.adjustment_refresh_timer.setInterval(250)
        self.adjustment_refresh_timer.timeout.connect(self.refresh_adjustment)

        self.summary_label = QLabel("选择 PDF 文件或文件夹后开始处理")
        self.summary_label.setObjectName("summaryLabel")
        self.summary_metrics = {
            "files": self.create_metric("文件", "0", "neutral"),
            "income": self.create_metric("收入", "0 / 0.00", "income"),
            "expense": self.create_metric("支出", "0 / 0.00", "expense"),
            "issues": self.create_metric("异常", "0", "warning"),
            "confidence": self.create_metric("可信度", "-", "neutral"),
        }
        self.adjust_preview_label = QLabel("启用调整后显示预览")
        self.adjust_preview_label.setObjectName("adjustmentStatusLabel")
        self.adjust_net_value = QLabel("0.00")
        self.adjust_net_value.setObjectName("adjustResultValue")
        self.adjust_check_value = QLabel("未启用")
        self.adjust_check_value.setObjectName("adjustResultValue")
        self.profit_base_value = QLabel("0.00")
        self.profit_base_value.setObjectName("profitResultValue")
        self.profit_generated_value = QLabel("0.00")
        self.profit_generated_value.setObjectName("profitResultValue")
        self.profit_check_value = QLabel("待处理")
        self.profit_check_value.setObjectName("profitResultValue")
        self.profit_check_value.setProperty("tone", "neutral")
        self.profit_hint_label = QLabel("处理流水后显示利润率校验")
        self.profit_hint_label.setObjectName("profitHintLabel")
        self.profit_hint_label.setWordWrap(True)
        self.drop_hint_label = QLabel("拖入 PDF/Excel 文件，或点击选择文件夹\n支持多文件同时导入")
        self.drop_hint_label.setObjectName("dropHint")
        self.drop_hint_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        add_folder = QPushButton("选择文件夹")
        clear = QPushButton("清空")
        run = QPushButton("开始处理")
        self.export_income_json_button = QPushButton("经营佐证")
        self.export_salary_json_button = QPushButton("工资佐证")
        run.setObjectName("primaryButton")
        self.export_income_json_button.setObjectName("exportButton")
        self.export_salary_json_button.setObjectName("exportButton")
        for button in [run, add_folder, clear, self.export_salary_json_button, self.export_income_json_button]:
            button.setFixedSize(112, 40)
        add_folder.clicked.connect(self.add_folder)
        clear.clicked.connect(self.clear)
        run.clicked.connect(self.run)
        self.export_income_json_button.clicked.connect(self.export_income_proof_json)
        self.export_salary_json_button.clicked.connect(self.export_salary_income_proof_json)

        self.date_filter = QCheckBox("筛选")
        self.date_filter.setChecked(True)
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("yyyy-MM-dd")
        self.start_date.setFixedWidth(156)
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("yyyy-MM-dd")
        self.end_date.setFixedWidth(156)
        start_date, end_date = default_recent_month_range()
        self.start_date.setDate(start_date)
        self.end_date.setDate(end_date)

        date_filter_panel = QFrame()
        date_filter_panel.setObjectName("inlineFilterPanel")
        date_filter_panel.setMinimumWidth(410)
        date_filter_panel.setMinimumHeight(44)
        date_filter_layout = QHBoxLayout(date_filter_panel)
        date_filter_layout.setContentsMargins(10, 6, 10, 6)
        date_filter_layout.setSpacing(8)
        date_filter_layout.addWidget(self.date_filter)
        date_filter_layout.addWidget(self.start_date)
        date_filter_layout.addWidget(self.end_date)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)
        toolbar.addWidget(run)
        toolbar.addWidget(add_folder)
        toolbar.addWidget(clear)
        toolbar.addWidget(self.export_salary_json_button)
        toolbar.addWidget(self.export_income_json_button)
        toolbar.addStretch(1)

        self.income_adjust = QPushButton("微信")
        self.income_adjust.setCheckable(True)
        self.balance_adjust = QPushButton("银行")
        self.balance_adjust.setCheckable(True)
        self.adjust_amount = QLineEdit()
        self.adjust_amount.setPlaceholderText("调整金额")
        self.adjust_amount.setFixedHeight(34)
        self.declared_month_income = QLineEdit()
        self.share_ratio = QLineEdit()
        self.share_ratio.setPlaceholderText("占股比例（%）")
        self.share_ratio.setText("100")
        self.profit_rate = QLineEdit()
        self.profit_rate.setPlaceholderText("%")
        self.profit_rate.setText("5")
        self.adjust_start_month = QDateEdit()
        self.adjust_start_month.setCalendarPopup(True)
        self.adjust_start_month.setDisplayFormat("yyyy-MM")
        self.adjust_start_month.setFixedWidth(108)
        self.adjust_end_month = QDateEdit()
        self.adjust_end_month.setCalendarPopup(True)
        self.adjust_end_month.setDisplayFormat("yyyy-MM")
        self.adjust_end_month.setFixedWidth(108)
        self.adjust_start_month.setDate(start_date)
        self.adjust_end_month.setDate(end_date)
        self.random_adjust = QCheckBox("固定分配")

        self.income_adjust.toggled.connect(self.on_adjustment_mode_changed)
        self.balance_adjust.toggled.connect(self.on_adjustment_mode_changed)
        self.random_adjust.stateChanged.connect(self.schedule_adjustment_refresh)
        self.adjust_amount.textChanged.connect(self.schedule_adjustment_refresh)
        self.adjust_start_month.dateChanged.connect(self.schedule_adjustment_refresh)
        self.adjust_end_month.dateChanged.connect(self.schedule_adjustment_refresh)
        self.declared_month_income.textChanged.connect(self.update_profit_preview)
        self.share_ratio.textChanged.connect(self.update_profit_preview)
        self.profit_rate.textChanged.connect(self.update_profit_preview)

        self.overview = DropTable()
        self.monthly = DropTable()
        self.details = DropTable()
        self.issue_table = DropTable()
        for table in (self.overview, self.monthly, self.details, self.issue_table):
            table.filesDropped.connect(self.add_paths)

        self.tabs = QTabWidget()
        self.monthly_tab_index = self.tabs.addTab(self.table_shell(self.monthly), "月度统计")
        self.tabs.addTab(self.table_shell(self.overview), "文件汇总")
        self.tabs.addTab(self.table_shell(self.details), "流水明细")
        self.tabs.addTab(self.table_shell(self.issue_table), "异常提示")

        central = DropWidget()
        central.filesDropped.connect(self.add_paths)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(10, 10, 10, 8)
        layout.setSpacing(10)
        layout.addLayout(toolbar)

        main_panel = QFrame()
        main_panel.setObjectName("card")
        main_layout = QVBoxLayout(main_panel)
        main_layout.setSpacing(10)
        metrics_bar = QFrame()
        metrics_bar.setObjectName("metricsBar")
        metrics_layout = QHBoxLayout(metrics_bar)
        metrics_layout.setContentsMargins(10, 8, 10, 8)
        metrics_layout.setSpacing(0)
        for index, metric in enumerate(self.summary_metrics.values()):
            metrics_layout.addWidget(metric, 1)
            if index < len(self.summary_metrics) - 1:
                divider = QFrame()
                divider.setObjectName("metricDivider")
                divider.setFrameShape(QFrame.Shape.VLine)
                metrics_layout.addWidget(divider)
        main_layout.addWidget(metrics_bar)
        self.summary_label.setVisible(False)
        main_layout.addWidget(self.tabs)

        left_area = QVBoxLayout()
        left_area.setSpacing(12)
        left_area.addWidget(main_panel, 1)

        side_content = QWidget()
        side_content.setFixedWidth(430)
        side_layout = QVBoxLayout(side_content)
        side_layout.setContentsMargins(0, 0, 0, 0)
        side_layout.setSpacing(14)
        side_layout.addWidget(date_filter_panel)

        income_calc_panel = QFrame()
        income_calc_panel.setObjectName("sideSection")
        income_calc_layout = QVBoxLayout(income_calc_panel)
        income_calc_layout.setContentsMargins(14, 14, 14, 14)
        income_calc_layout.setSpacing(10)
        self.side_panel = income_calc_panel
        self.side_panel_layout = side_layout

        adjustment_section_title = QLabel("流水调整")
        adjustment_section_title.setObjectName("sectionTitle")
        income_calc_layout.addWidget(adjustment_section_title)
        self.income_adjust.setObjectName("adjustModeOption")
        self.balance_adjust.setObjectName("adjustModeOption")
        adjust_option_row = QHBoxLayout()
        adjust_option_row.setContentsMargins(0, 0, 0, 0)
        adjust_option_row.setSpacing(10)
        adjust_option_row.addWidget(self.adjust_amount, 1)
        adjust_option_row.addWidget(self.income_adjust, 0)
        adjust_option_row.addWidget(self.balance_adjust, 0)
        income_calc_layout.addLayout(adjust_option_row)
        preview_card = QFrame()
        preview_card.setObjectName("adjustResultCard")
        preview_layout = QVBoxLayout(preview_card)
        preview_layout.setContentsMargins(10, 6, 10, 7)
        preview_layout.setSpacing(4)
        result_row = QHBoxLayout()
        net_box = QVBoxLayout()
        net_label = QLabel("调整后收支差额")
        net_label.setObjectName("compactResultLabel")
        net_box.addWidget(net_label)
        self.adjust_net_value.setObjectName("compactResultValue")
        net_box.addWidget(self.adjust_net_value)
        check_box = QVBoxLayout()
        check_label = QLabel("平衡校验")
        check_label.setObjectName("compactResultLabel")
        check_box.addWidget(check_label)
        self.adjust_check_value.setObjectName("compactResultValue")
        check_box.addWidget(self.adjust_check_value)
        result_row.addLayout(net_box)
        divider = QFrame()
        divider.setObjectName("metricDivider")
        divider.setFrameShape(QFrame.Shape.VLine)
        result_row.addWidget(divider)
        result_row.addLayout(check_box)
        preview_layout.addLayout(result_row)
        preview_layout.addWidget(self.adjust_preview_label)
        income_calc_layout.addWidget(preview_card)

        profit_title = QLabel("收入测算")
        profit_title.setObjectName("sectionTitle")
        declared_label = QLabel("系统月收入")
        declared_label.setObjectName("declaredIncomeLabel")
        rate_label = QLabel("利润率（%）")
        rate_label.setObjectName("profitInputLabel")
        share_label = QLabel("占股比例（%）")
        share_label.setObjectName("profitInputLabel")
        profit_card = QFrame()
        profit_card.setObjectName("adjustResultCard")
        profit_layout = QVBoxLayout(profit_card)
        profit_layout.setContentsMargins(10, 8, 10, 8)
        profit_layout.setSpacing(6)
        base_label = QLabel("月均流水收入（固定÷6）")
        base_label.setObjectName("profitCardLabel")
        generated_label = QLabel("占股与利润率后佐证收入")
        generated_label.setObjectName("profitCardLabel")
        status_label = QLabel("校验")
        status_label.setObjectName("profitCardLabel")
        self.profit_base_value.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.profit_generated_value.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.profit_check_value.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        for label, value in (
            (base_label, self.profit_base_value),
            (generated_label, self.profit_generated_value),
            (status_label, self.profit_check_value),
        ):
            row = QHBoxLayout()
            row.addWidget(label, 1)
            row.addWidget(value, 0)
            profit_layout.addLayout(row)
        income_calc_layout.addWidget(profit_title)
        profit_inputs_row = QHBoxLayout()
        profit_inputs_row.setSpacing(10)
        declared_box = QVBoxLayout()
        declared_box.setSpacing(5)
        declared_box.addWidget(declared_label)
        declared_box.addWidget(self.declared_month_income)
        share_box = QVBoxLayout()
        share_box.setSpacing(5)
        share_box.addWidget(share_label)
        share_box.addWidget(self.share_ratio)
        rate_box = QVBoxLayout()
        rate_box.setSpacing(5)
        rate_box.addWidget(rate_label)
        rate_box.addWidget(self.profit_rate)
        self.share_ratio.setFixedWidth(92)
        self.profit_rate.setFixedWidth(92)
        self.income_adjust.setFixedWidth(92)
        self.balance_adjust.setFixedWidth(92)
        profit_inputs_row.addLayout(declared_box, 1)
        profit_inputs_row.addLayout(share_box, 0)
        profit_inputs_row.addLayout(rate_box, 0)
        income_calc_layout.addLayout(profit_inputs_row)
        profit_layout.addWidget(self.profit_hint_label)
        income_calc_layout.addWidget(profit_card)
        side_layout.addWidget(income_calc_panel)

        content = QHBoxLayout()
        content.setSpacing(10)
        content.addLayout(left_area, 1)
        side_scroll = QScrollArea()
        side_scroll.setObjectName("sideScroll")
        side_scroll.setWidgetResizable(True)
        side_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        side_scroll.setWidget(side_content)
        side_scroll.setFixedWidth(444)
        side_column = QVBoxLayout()
        side_column.addWidget(side_scroll)
        content.addLayout(side_column)
        layout.addLayout(content, 1)
        self.setCentralWidget(central)
        self.setStatusBar(QStatusBar())

        self._apply_style()
        self.render_empty()

    def table_shell(self, table: QTableWidget) -> QFrame:
        shell = RoundedTableShell()
        shell.setObjectName("FlowTableShell")
        layout = QVBoxLayout(shell)
        layout.setContentsMargins(1, 1, 1, 1)
        layout.setSpacing(0)
        layout.addWidget(table)
        shell.set_corner_colors("#ffffff", "#dbe8fb")
        return shell

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        paths = [Path(url.toLocalFile()) for url in event.mimeData().urls()]
        self.add_paths(paths)

    def _apply_style(self):
        self.setStyleSheet(
            """
            QMainWindow { background: #f5f7fb; }
            QPushButton {
                min-height: 34px;
                padding: 0 14px;
                border: 1px solid #d5dbe5;
                border-radius: 12px;
                background: #ffffff;
                color: #263243;
            }
            QPushButton:hover { background: #eef3f8; }
            QPushButton:disabled {
                background: #f1f4f8;
                color: #9aa6b2;
                border-color: #dfe6f1;
            }
            QPushButton#primaryButton {
                background: #1769e0;
                color: #ffffff;
                border-color: #1769e0;
                font-weight: 700;
                padding: 0 16px;
            }
            QPushButton#primaryButton:hover { background: #0f5ccc; }
            QPushButton#exportButton {
                background: #ffffff;
                color: #1769e0;
                border-color: #9bbff2;
                font-size: 14px;
                font-weight: 500;
            }
            QFrame#card, QFrame#sidePanel, QFrame#sideSection {
                background: #ffffff;
                border: 1px solid #e3e8f0;
                border-radius: 18px;
            }
            QFrame#sidePanel, QFrame#sideSection { background: #fbfcfe; }
            QScrollArea#sideScroll {
                background: transparent;
                border: 0;
            }
            QScrollArea#sideScroll > QWidget > QWidget {
                background: transparent;
            }
            QFrame#inlineFilterPanel {
                background: #ffffff;
                border: 1px solid #dfe6f1;
                border-radius: 12px;
                min-height: 42px;
            }
            QFrame#metricsBar {
                background: #ffffff;
                border: 1px solid #dfe6f1;
                border-radius: 12px;
            }
            QFrame#metricItem { border: 0; background: transparent; }
            QFrame#metricDivider {
                color: #e1e7f0;
                background: #e1e7f0;
                max-width: 1px;
            }
            QLabel#metricLabel {
                color: #5f6b7a;
                font-size: 12px;
                font-weight: 500;
            }
            QLabel#metricValue {
                color: #162033;
                font-size: 20px;
                font-weight: 600;
            }
            QLabel#metricValue[tone="income"] { color: #138a4b; }
            QLabel#metricValue[tone="expense"] { color: #d93f2f; }
            QLabel#metricValue[tone="income"], QLabel#metricValue[tone="expense"] {
                font-size: 18px;
            }
            QLabel#metricValue[tone="warning"] { color: #e07a1f; }
            QLabel#metricValue[tone="confidenceVeryHigh"] { color: #0b6f3a; }
            QLabel#metricValue[tone="confidenceHigh"] { color: #138a4b; }
            QLabel#metricValue[tone="confidenceMedium"] { color: #e07a1f; }
            QLabel#metricValue[tone="confidenceLow"] { color: #d93f2f; }
            QLabel#metricValue[tone="neutral"] { color: #162033; }
            QLabel {
                font-family: "Microsoft YaHei UI", "Segoe UI", Arial;
                color: #162033;
            }
            QLabel:disabled {
                color: #9aa6b2;
            }
            QLabel#cardTitle {
                font-size: 16px;
                font-weight: 700;
                color: #1f2d3d;
            }
            QLabel#fieldLabel {
                color: #5f6b7a;
                font-weight: 600;
            }
            QLabel#dropHint {
                min-height: 64px;
                color: #4b5b6d;
                border: 1px dashed #9bbff2;
                border-radius: 12px;
                background: #f8fbff;
                font-size: 14px;
                font-weight: 400;
            }
            QFrame#FlowTableShell {
                background: #ffffff;
                border: 1px solid #dbe8fb;
                border-radius: 12px;
            }
            QTableWidget {
                gridline-color: transparent;
                selection-background-color: #cfe5ff;
                selection-color: #162033;
                background: transparent;
                alternate-background-color: #f8fbff;
                color: #162033;
                border: 0;
                border-radius: 0;
                padding: 0;
            }
            QTableWidget::viewport {
                background: transparent;
                border: 0;
                border-radius: 0;
            }
            QAbstractScrollArea::corner,
            QTableCornerButton::section {
                background: #f0f3f7;
                border: 0;
                border-radius: 0;
            }
            QTableWidget::item {
                color: #162033;
                background: transparent;
            }
            QTableWidget::item:selected {
                color: #162033;
                background: #cfe5ff;
            }
            QTableWidget::item:disabled {
                color: #9aa6b2;
            }
            QHeaderView::section {
                background: #f0f3f7;
                padding: 7px;
                border: 0;
                font-family: "Microsoft YaHei UI", "Segoe UI", Arial;
                font-size: 13px;
                font-weight: 500;
                color: #1f2d3d;
            }
            QHeaderView {
                background: #f0f3f7;
                border: 0;
                border-radius: 0;
            }
            QTabWidget::pane { border: 0; top: -1px; }
            QTabBar::tab {
                background: transparent;
                border: 0;
                border-bottom: 3px solid transparent;
                padding: 10px 20px;
                margin-right: 8px;
                color: #5f6b7a;
                font-size: 15px;
                font-weight: 400;
            }
            QTabBar::tab:selected {
                color: #1769e0;
                font-weight: 500;
                border-bottom-color: #1769e0;
            }
            QLineEdit, QDateEdit {
                min-height: 28px;
                padding: 0 8px;
                border: 1px solid #d5dbe5;
                border-radius: 10px;
                background: #ffffff;
                color: #162033;
                selection-background-color: #cfe5ff;
                selection-color: #162033;
            }
            QLineEdit:disabled, QDateEdit:disabled {
                background: #f1f4f8;
                color: #9aa6b2;
                border-color: #dfe6f1;
            }
            QDateEdit {
                padding: 0 20px 0 8px;
                font-size: 12px;
            }
            QCheckBox {
                color: #263243;
                spacing: 8px;
                background: transparent;
            }
            QFrame#inlineFilterPanel QCheckBox {
                background: transparent;
                border: 0;
                padding: 0;
                font-weight: 400;
            }
            QPushButton#adjustModeOption {
                min-height: 32px;
                border-radius: 10px;
                border: 1px solid #14b981;
                background: transparent;
                color: #14b981;
                font-size: 13px;
                font-weight: 600;
                padding: 0;
            }
            QPushButton#adjustModeOption:hover {
                background: rgba(20, 185, 129, 0.12);
                border-color: #20d99a;
                color: #20d99a;
            }
            QPushButton#adjustModeOption:checked {
                background: rgba(20, 185, 129, 0.16);
                border-color: #20d99a;
                color: #20d99a;
            }
            QCheckBox:disabled {
                color: #9aa6b2;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border-radius: 4px;
                border: 1px solid #cfd7e3;
                background: #ffffff;
            }
            QCheckBox::indicator:checked {
                background: #4cc2ff;
                border-color: #4cc2ff;
                image: url(assets/check.svg);
            }
            QCheckBox::indicator:disabled {
                background: #e1e5ea;
                border-color: #d5dbe5;
            }
            QDateEdit::drop-down {
                width: 14px;
                border: 0;
                background: transparent;
                subcontrol-origin: padding;
                subcontrol-position: top right;
            }
            QDateEdit::down-arrow {
                image: url(assets/down_arrow.svg);
                width: 8px;
                height: 8px;
                margin-right: 4px;
            }
            QLabel#summaryLabel {
                padding: 10px 12px;
                background: #f8fbff;
                border: 1px solid #dbe8fb;
                border-radius: 12px;
                color: #1f2d3d;
            }
            QLabel#adjustPreviewLabel {
                color: #5f6b7a;
            }
            QLabel#adjustmentStatusLabel,
            QLabel#profitHintLabel {
                color: #5f6b7a;
                font-size: 12px;
                font-weight: 400;
                padding-top: 4px;
            }
            QLabel#declaredIncomeLabel {
                color: #5f6b7a;
                font-size: 12px;
                font-weight: 400;
                margin-left: 8px;
            }
            QLabel#profitCardLabel {
                color: #5f6b7a;
                font-size: 12px;
                font-weight: 400;
            }
            QLabel#profitInputLabel {
                color: #5f6b7a;
                font-size: 12px;
                font-weight: 400;
            }
            QLabel#sectionTitle {
                color: #263243;
                font-size: 16px;
                font-weight: 600;
                padding-top: 4px;
                padding-bottom: 2px;
            }
            QLabel#sectionHintLabel {
                color: #5f6b7a;
                font-size: 11px;
                font-weight: 400;
            }
            QFrame#adjustResultCard {
                padding: 8px;
                background: #f8fbff;
                border: 1px solid #dbe8fb;
                border-radius: 12px;
            }
            QLabel#adjustResultValue {
                color: #1f9d55;
                font-size: 18px;
                font-weight: 600;
            }
            QLabel#compactResultLabel {
                color: #5f6b7a;
                font-size: 11px;
                font-weight: 400;
            }
            QLabel#compactResultValue {
                color: #1f9d55;
                font-size: 13px;
                font-weight: 400;
            }
            QLabel#profitResultValue {
                color: #162033;
                font-size: 16px;
                font-weight: 600;
            }
            QLabel#profitResultValue[tone="ok"] { color: #138a4b; }
            QLabel#profitResultValue[tone="warning"] { color: #d93f2f; }
            QLabel#profitResultValue[tone="neutral"] { color: #162033; }
            QFrame#adjustResultCard QLabel#fieldLabel {
                font-size: 12px;
                font-weight: 500;
            }
            """
        )
        self._normalize_compact_control_heights()

    def apply_embedded_theme(self, colors: dict[str, str], dark: bool = False) -> None:
        self._embedded_theme_colors = dict(colors)
        self._embedded_theme_dark = dark
        self.setStyleSheet(self._embedded_theme_style(colors))
        for shell in self.findChildren(RoundedTableShell):
            shell.set_corner_colors(colors["card"], colors["border"])
        self._apply_table_scrollbar_theme(colors)
        self._normalize_compact_control_heights()

    def _normalize_compact_control_heights(self) -> None:
        for widget in (
            self.adjust_amount,
            self.income_adjust,
            self.balance_adjust,
            self.declared_month_income,
            self.share_ratio,
            self.profit_rate,
        ):
            widget.setFixedHeight(34)
        for button in self.findChildren(QPushButton):
            if button.objectName() in {"proofActionButton", "proofPrimaryActionButton"}:
                button.setFixedHeight(34)

    def _theme_asset_path(self, name: str) -> str:
        return (runtime_dir() / "assets" / name).as_posix()

    def _apply_table_scrollbar_theme(self, colors: dict[str, str]) -> None:
        scrollbar_style = f"""
        QScrollBar:horizontal {{
            background: {colors["table_base"]};
            border: 0;
            height: 8px;
            margin: 0;
        }}
        QScrollBar::handle:horizontal {{
            background: {colors["border2"]};
            border-radius: 4px;
            min-width: 36px;
        }}
        QScrollBar::handle:horizontal:hover {{
            background: {colors["muted"]};
        }}
        QScrollBar::add-line:horizontal,
        QScrollBar::sub-line:horizontal,
        QScrollBar::add-page:horizontal,
        QScrollBar::sub-page:horizontal {{
            background: {colors["table_base"]};
            border: 0;
            width: 0;
        }}
        """
        for table in (self.monthly, self.overview, self.details, self.issue_table):
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
            table.horizontalScrollBar().setStyleSheet(scrollbar_style)

    def _embedded_theme_style(self, colors: dict[str, str]) -> str:
        check_icon = self._theme_asset_path("check.svg")
        arrow_icon = self._theme_asset_path("down_arrow.svg")
        return f"""
        QMainWindow, QWidget {{
            background: {colors["bg"]};
            color: {colors["text"]};
            font-family: "Microsoft YaHei UI";
            font-size: 14px;
        }}
        QPushButton {{
            min-height: 30px;
            padding: 0 11px;
            border: 1px solid {colors["border2"]};
            border-radius: 12px;
            background: {colors["field"]};
            color: {colors["text"]};
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton:hover {{
            background: {colors["soft"]};
            border-color: {colors["muted"]};
        }}
        QPushButton:disabled {{
            background: {colors["soft"]};
            color: {colors["disabled"]};
            border-color: {colors["border"]};
        }}
        QPushButton#primaryButton {{
            min-height: 34px;
            padding: 0 14px;
            border-radius: 12px;
            background: {colors["field"]};
            border: 1px solid {colors["accent"]};
            color: {colors["accent"]};
            font-weight: 700;
        }}
        QPushButton#primaryButton:hover {{
            background: {colors["selection"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent_hover"]};
        }}
        QPushButton#exportButton {{
            min-height: 30px;
            padding: 0 12px;
            border-radius: 12px;
            background: {colors["soft"]};
            border-color: {colors["accent"]};
            color: {colors["accent"]};
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton#exportButton:hover {{
            background: {colors["selection"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent"]};
        }}
        QFrame#card, QFrame#sidePanel, QFrame#sideSection {{
            background: {colors["card"]};
            border: 1px solid {colors["border"]};
            border-radius: 18px;
        }}
        QFrame#card:hover, QFrame#sidePanel:hover, QFrame#sideSection:hover {{
            border-color: {colors["accent"]};
        }}
        QFrame#sidePanel, QFrame#sideSection {{
            background: {colors["panel"]};
        }}
        QScrollArea#sideScroll,
        QScrollArea#sideScroll > QWidget > QWidget {{
            background: transparent;
            border: 0;
        }}
        QFrame#inlineFilterPanel, QFrame#metricsBar, QFrame#adjustResultCard, QFrame#incomeProofPanel {{
            background: {colors["soft"]};
            border: 1px solid {colors["border"]};
            border-radius: 12px;
        }}
        QFrame#incomeProofPanel {{
            border-radius: 18px;
        }}
        QFrame#inlineFilterPanel:hover, QFrame#metricsBar:hover,
        QFrame#adjustResultCard:hover, QFrame#incomeProofPanel:hover {{
            border-color: {colors["accent"]};
        }}
        QFrame#inlineFilterPanel {{
            min-height: 42px;
        }}
        QFrame#metricItem {{
            border: 0;
            background: transparent;
        }}
        QFrame#metricDivider {{
            color: {colors["border"]};
            background: {colors["border"]};
            max-width: 1px;
        }}
        QLabel {{
            background: transparent;
            color: {colors["text"]};
        }}
        QLabel:disabled {{
            color: {colors["disabled"]};
        }}
        QLabel#cardTitle {{
            color: {colors["text"]};
            font-size: 16px;
            font-weight: 600;
        }}
        QLabel#metricLabel, QLabel#fieldLabel, QLabel#adjustPreviewLabel {{
            color: {colors["muted"]};
            font-weight: 600;
        }}
        QLabel#adjustmentStatusLabel,
        QLabel#profitHintLabel,
        QLabel#IncomeProofStatus {{
            color: {colors["muted"]};
            font-size: 12px;
            font-weight: 400;
            padding-top: 4px;
        }}
        QLabel#FlowBindingStatus {{
            color: {colors["muted"]};
            font-size: 11px;
            font-weight: 400;
            padding: 0;
        }}
        QLabel#IncomeProofStatus[tone="warning"] {{
            color: {colors["bad"]};
        }}
        QLabel#declaredIncomeLabel {{
            color: {colors["muted"]};
            font-size: 12px;
            font-weight: 400;
            margin-left: 8px;
        }}
        QLabel#profitCardLabel {{
            color: {colors["muted"]};
            font-size: 12px;
            font-weight: 400;
        }}
        QLabel#profitInputLabel {{
            color: {colors["muted"]};
            font-size: 12px;
            font-weight: 400;
        }}
        QLabel#sectionTitle {{
            color: {colors["text"]};
            font-size: 16px;
            font-weight: 600;
            padding-top: 4px;
            padding-bottom: 2px;
        }}
        QLabel#sectionHintLabel {{
            color: {colors["muted"]};
            font-size: 11px;
            font-weight: 400;
        }}
        QLabel#metricValue {{
            color: {colors["text"]};
            font-size: 17px;
            font-weight: 600;
        }}
        QLabel#metricValue[tone="income"],
        QLabel#metricValue[tone="confidenceVeryHigh"],
        QLabel#metricValue[tone="confidenceHigh"],
        QLabel#profitResultValue[tone="ok"] {{
            color: {colors["good"]};
        }}
        QLabel#metricValue[tone="expense"],
        QLabel#metricValue[tone="confidenceLow"],
        QLabel#profitResultValue[tone="warning"] {{
            color: {colors["bad"]};
        }}
        QLabel#metricValue[tone="warning"],
        QLabel#metricValue[tone="confidenceMedium"] {{
            color: {colors["accent"]};
        }}
        QLabel#summaryLabel {{
            padding: 8px 10px;
            background: {colors["soft"]};
            border: 1px solid {colors["border"]};
            border-radius: 12px;
            color: {colors["text"]};
        }}
        QLabel#dropHint {{
            min-height: 56px;
            color: {colors["muted"]};
            border: 1px dashed {colors["accent"]};
            border-radius: 12px;
            background: {colors["drop_bg"]};
            font-size: 13px;
            font-weight: 400;
        }}
        QFrame#FlowTableShell {{
            background: {colors["table_base"]};
            border: 1px solid {colors["border"]};
            border-radius: 12px;
        }}
        QTableWidget {{
            gridline-color: transparent;
            selection-background-color: {colors["selection"]};
            selection-color: {colors["selection_text"]};
            background: transparent;
            alternate-background-color: {colors["table_alt"]};
            color: {colors["text"]};
            border: 0;
            border-radius: 0;
            padding: 0;
        }}
        QTableWidget::viewport {{
            background: transparent;
            border: 0;
            border-radius: 0;
        }}
        QAbstractScrollArea::corner,
        QTableCornerButton::section {{
            background: {colors["soft"]};
            border: 0;
            border-radius: 0;
        }}
        QTableWidget::item {{
            color: {colors["text"]};
            background: transparent;
        }}
        QTableWidget::item:selected {{
            color: {colors["selection_text"]};
            background: {colors["selection"]};
        }}
        QHeaderView::section {{
            background: {colors["soft"]};
            padding: 6px 8px;
            border: 0;
            color: {colors["text"]};
            font-weight: 600;
        }}
        QHeaderView {{
            background: {colors["soft"]};
            border: 0;
            border-radius: 0;
        }}
        QTabWidget::pane {{
            border: 0;
            top: -1px;
        }}
        QTabBar::tab {{
            background: transparent;
            border: 0;
            border-bottom: 4px solid transparent;
            border-radius: 2px;
            padding: 8px 13px 7px 13px;
            margin-right: 12px;
            color: {colors["muted"]};
            font-size: 13px;
            font-weight: 400;
        }}
        QTabBar::tab:selected {{
            color: {colors["accent"]};
            font-weight: 600;
            border-bottom-color: {colors["accent"]};
        }}
        QWidget#FlowDetailTabNav {{
            background: {colors["card"]};
        }}
        QStatusBar {{
            min-height: 22px;
            max-height: 22px;
            padding: 0 8px;
            background: {colors["status_bg"]};
            color: {colors["muted"]};
            border: 0;
            font-size: 12px;
        }}
        QStatusBar::item {{
            border: 0;
        }}
        QStatusBar QLabel {{
            background: transparent;
            color: {colors["muted"]};
        }}
        QWidget#FlowDetailTabIndicatorSlot {{
            background: transparent;
            border: 0;
        }}
        QPushButton#FlowDetailTabButton {{
            min-height: 32px;
            max-height: 32px;
            padding: 0;
            border: 0;
            border-radius: 0;
            background: transparent;
            color: {colors["muted"]};
            font-size: 15px;
            font-weight: 500;
        }}
        QPushButton#FlowDetailTabButton:hover {{
            background: transparent;
            color: {colors["accent_hover"]};
        }}
        QPushButton#FlowDetailTabButton:checked {{
            background: transparent;
            color: {colors["accent"]};
            font-weight: 700;
        }}
        QFrame#FlowDetailTabIndicator {{
            background: transparent;
            border: 0;
            border-radius: 2px;
        }}
        QFrame#FlowDetailTabIndicator[active="true"] {{
            background: {colors["accent"]};
        }}
        QLineEdit, QDateEdit {{
            min-height: 28px;
            min-width: 96px;
            padding: 0 8px;
            border: 1px solid {colors["border2"]};
            border-radius: 10px;
            background: {colors["field"]};
            color: {colors["text"]};
            font-size: 13px;
            selection-background-color: {colors["selection"]};
            selection-color: {colors["selection_text"]};
        }}
        QLineEdit {{
            min-width: 0;
        }}
        QDateEdit {{
            padding: 0 20px 0 8px;
            font-size: 12px;
        }}
        QDateEdit::drop-down {{
            width: 14px;
            border: 0;
            background: transparent;
            subcontrol-origin: padding;
            subcontrol-position: top right;
        }}
        QDateEdit::down-arrow {{
            image: url({arrow_icon});
            width: 8px;
            height: 8px;
            margin-right: 4px;
        }}
        QLineEdit:focus, QDateEdit:focus {{
            border-color: {colors["accent"]};
        }}
        QLineEdit:disabled, QDateEdit:disabled {{
            background: {colors["soft"]};
            color: {colors["disabled"]};
            border-color: {colors["border"]};
        }}
        QPlainTextEdit#IncomeProofReview {{
            padding: 8px;
            border: 1px solid {colors["border2"]};
            border-radius: 10px;
            background: {colors["field"]};
            color: {colors["text"]};
            font-size: 12px;
            selection-background-color: {colors["selection"]};
            selection-color: {colors["selection_text"]};
        }}
        QPlainTextEdit#IncomeProofReview:focus {{
            border-color: {colors["accent"]};
        }}
        QPushButton#proofPrimaryButton {{
            min-height: 30px;
            border-radius: 10px;
            background: {colors["field"]};
            border: 1px solid {colors["accent"]};
            color: {colors["accent"]};
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton#proofPrimaryButton:hover {{
            background: {colors["selection"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent_hover"]};
        }}
        QPushButton#proofSecondaryButton {{
            min-height: 30px;
            border-radius: 10px;
            background: {colors["field"]};
            border: 1px solid {colors["border2"]};
            color: {colors["text"]};
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton#proofSecondaryButton:hover {{
            background: {colors["soft"]};
            border-color: {colors["accent"]};
            color: {colors["accent"]};
        }}
        QPushButton#proofActionButton {{
            min-height: 30px;
            border-radius: 10px;
            background: {colors["field"]};
            border: 1px solid {colors["border2"]};
            color: {colors["text"]};
            font-size: 13px;
            font-weight: 400;
        }}
        QPushButton#proofActionButton:hover {{
            background: {colors["soft"]};
            border-color: {colors["accent"]};
            color: {colors["text"]};
        }}
        QPushButton#proofActionButton:pressed {{
            background: {colors["selection"]};
            border-color: {colors["accent_hover"]};
            color: {colors["selection_text"]};
        }}
        QPushButton#proofPrimaryActionButton {{
            min-height: 30px;
            border-radius: 10px;
            background: {colors["accent"]};
            border: 1px solid {colors["accent"]};
            color: {colors["accent_text"]};
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton#proofPrimaryActionButton:hover {{
            background: {colors["accent_hover"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent_text"]};
        }}
        QPushButton#proofPrimaryActionButton:pressed {{
            background: {colors["checked_fill"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent_text"]};
        }}
        QCheckBox {{
            color: {colors["text"]};
            spacing: 6px;
            font-size: 13px;
            background: transparent;
        }}
        QFrame#inlineFilterPanel QCheckBox {{
            background: transparent;
            border: 0;
            padding: 0;
            font-size: 13px;
            font-weight: 400;
        }}
        QPushButton#adjustModeOption {{
            min-height: 32px;
            border-radius: 10px;
            border: 1px solid {colors["accent"]};
            background: transparent;
            color: {colors["accent"]};
            font-size: 13px;
            font-weight: 600;
            padding: 0;
        }}
        QPushButton#adjustModeOption:hover {{
            background: {colors["selection"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent_hover"]};
        }}
        QPushButton#adjustModeOption:checked {{
            background: {colors["accent"]};
            border-color: {colors["accent"]};
            color: {colors["accent_text"]};
        }}
        QPushButton#adjustModeOption:checked:hover {{
            background: {colors["accent_hover"]};
            border-color: {colors["accent_hover"]};
            color: {colors["accent_text"]};
        }}
        QCheckBox#proofOption {{
            min-height: 28px;
            padding: 0;
            border: 0;
            border-radius: 0;
            background: transparent;
            color: {colors["text"]};
            font-size: 12px;
            font-weight: 400;
        }}
        QCheckBox#proofOption:hover {{
            background: transparent;
            color: {colors["text"]};
        }}
        QCheckBox#proofOption:checked {{
            background: transparent;
            color: {colors["text"]};
        }}
        QCheckBox#proofOption:checked:hover {{
            background: transparent;
            color: {colors["text"]};
        }}
        QPushButton#proofConfirmOption {{
            min-height: 30px;
            padding: 0 10px;
            border: 1px solid {colors["accent"]};
            border-radius: 10px;
            background: {colors["field"]};
            color: {colors["text"]};
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton#proofConfirmOption:hover {{
            background: {colors["selection"]};
            border-color: {colors["accent_hover"]};
            color: {colors["text"]};
        }}
        QPushButton#proofConfirmOption:checked {{
            border-color: {colors["accent"]};
            background: {colors["checked_fill"]};
            color: {colors["checked_text"]};
        }}
        QPushButton#proofConfirmOption:checked:hover {{
            border-color: {colors["accent_hover"]};
            background: {colors["checked_hover"]};
            color: {colors["checked_hover_text"]};
        }}
        QCheckBox:disabled {{
            color: {colors["disabled"]};
        }}
        QCheckBox::indicator {{
            width: 13px;
            height: 13px;
            border-radius: 4px;
            border: 1px solid {colors["border2"]};
            background: {colors["field"]};
        }}
        QCheckBox::indicator:checked {{
            background: {colors["accent"]};
            border-color: {colors["accent"]};
            image: url({check_icon});
        }}
        QCheckBox#proofOption::indicator {{
            width: 14px;
            height: 14px;
            border-radius: 5px;
            border: 1px solid {colors["border2"]};
            background: {colors["field"]};
        }}
        QCheckBox#proofOption::indicator:checked {{
            background: {colors["accent"]};
            border-color: {colors["accent"]};
            image: url({check_icon});
        }}
        QLabel#adjustResultValue,
        QLabel#profitResultValue {{
            color: {colors["text"]};
            font-weight: 600;
        }}
        QLabel#adjustResultValue {{
            color: {colors["good"]};
            font-size: 15px;
        }}
        QLabel#profitResultValue {{
            color: {colors["text"]};
            font-size: 13px;
        }}
        QLabel#compactResultLabel {{
            color: {colors["muted"]};
            font-size: 11px;
            font-weight: 400;
        }}
        QLabel#compactResultValue {{
            color: {colors["good"]};
            font-size: 13px;
            font-weight: 400;
        }}
        """

    def add_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "选择银行流水文件", "", "支持的文件 (*.pdf *.xlsx *.xlsm);;PDF 文件 (*.pdf);;Excel 文件 (*.xlsx *.xlsm)")
        self.add_paths([Path(file) for file in files])

    def add_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择包含 PDF 的文件夹")
        if folder:
            self.add_paths([Path(folder)])

    def add_paths(self, paths: list[Path]):
        inputs: list[Path] = []
        for path in paths:
            if path.is_dir():
                for suffix in SUPPORTED_INPUTS:
                    inputs.extend(sorted(path.rglob(f"*{suffix}")))
            elif path.suffix.lower() in SUPPORTED_INPUTS:
                inputs.append(path)
        self.paths = []
        self.results = []
        self.issues = []
        known = set()
        for input_path in inputs:
            resolved = input_path.resolve()
            if resolved not in known:
                self.paths.append(input_path)
                known.add(resolved)
        self.render_selected()

    def clear(self):
        self.paths = []
        self.results = []
        self.issues = []
        self.adjustment_result = AdjustmentResult()
        self.render_empty()

    def create_metric(self, label: str, value: str, tone: str) -> QFrame:
        frame = QFrame()
        frame.setObjectName("metricItem")
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(8, 3, 8, 3)
        title = QLabel(label)
        title.setObjectName("metricLabel")
        number = QLabel(value)
        number.setObjectName("metricValue")
        number.setProperty("tone", tone)
        number.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(title)
        layout.addWidget(number)
        frame.value_label = number
        return frame

    def set_metric(self, key: str, value: str, tone: str | None = None):
        metric = self.summary_metrics.get(key)
        if metric is not None:
            if tone is not None:
                metric.value_label.setProperty("tone", tone)
                metric.value_label.style().unpolish(metric.value_label)
                metric.value_label.style().polish(metric.value_label)
            metric.value_label.setText(self.format_metric_value(key, value))

    def format_metric_value(self, key: str, value: str) -> str:
        if key not in ("income", "expense") or " / " not in value:
            return value
        count, amount = value.split(" / ", 1)
        return f'<span style="font-weight:700;">{count}</span><span style="font-weight:400; font-size:13px;"> / {amount}</span>'

    def run(self):
        if not self.paths:
            QMessageBox.information(self, "提示", "请先选择 PDF/Excel 文件或文件夹。")
            return
        start_date, end_date = self.selected_date_range()
        if start_date is not None and end_date is not None and start_date > end_date:
            QMessageBox.warning(self, "日期范围错误", "开始日期不能晚于结束日期。")
            return
        pdf_passwords = self.collect_pdf_passwords()
        if pdf_passwords is None:
            return
        self.worker = Worker(self.paths, start_date, end_date, pdf_passwords)
        self.worker.progress.connect(self.statusBar().showMessage)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()

    def collect_pdf_passwords(self) -> dict[Path, str] | None:
        from bankflow_v2.pdf_password import pdf_requires_password, validate_pdf_password

        passwords: dict[Path, str] = {}
        last_password = ""
        for path in self.paths:
            if path.suffix.lower() != ".pdf":
                continue
            try:
                needs_password = pdf_requires_password(path)
            except Exception:
                needs_password = False
            if not needs_password:
                if not self.should_try_pdf_password_for_unrecognized(path):
                    continue
            if last_password and validate_pdf_password(path, last_password):
                passwords[path] = last_password
                continue
            while True:
                password, ok = QInputDialog.getText(
                    self,
                    "PDF密码",
                    f"{path.name} 需要密码，请输入后继续解析。",
                    QLineEdit.EchoMode.Password,
                    last_password,
                )
                if not ok:
                    self.statusBar().showMessage("已取消处理加密 PDF")
                    return None
                if validate_pdf_password(path, password):
                    passwords[path] = password
                    last_password = password
                    break
                QMessageBox.warning(self, "密码错误", f"{path.name} 密码不正确，请重新输入。")
        return passwords

    def should_try_pdf_password_for_unrecognized(self, path: Path) -> bool:
        try:
            from bankflow_v2.auto_detect import detect_bank_type

            detection = detect_bank_type(str(path))
        except Exception:
            detection = None
        if detection is not None and getattr(detection, "bank_id", ""):
            return False
        reply = QMessageBox.question(
            self,
            "尝试PDF密码",
            f"{path.name} 当前无法识别银行。\n如果这是带密码的 PDF，是否输入密码后重试解析？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        return reply == QMessageBox.StandardButton.Yes

    def selected_date_range(self) -> tuple[datetime | None, datetime | None]:
        if not self.date_filter.isChecked():
            return None, None
        start = self.start_date.date().toPyDate()
        end = self.end_date.date().toPyDate()
        return datetime.combine(start, time.min), datetime.combine(end, time.max)

    def on_finished(self, results: list[FileResult], issues: list[Issue]):
        self.results = results
        self.issues = issues
        self.render_results()
        self.statusBar().showMessage("处理完成")

    def on_adjustment_mode_changed(self, *_args):
        sender = self.sender()
        if sender is self.income_adjust and self.income_adjust.isChecked() and self.balance_adjust.isChecked():
            self.balance_adjust.setChecked(False)
        elif sender is self.balance_adjust and self.balance_adjust.isChecked() and self.income_adjust.isChecked():
            self.income_adjust.setChecked(False)
        self.refresh_adjustment()

    def refresh_adjustment(self, *_args):
        if self.results:
            configs = self.adjustment_configs()
            self.render_adjustment_only(configs)
            if any(config.enabled for config in configs):
                self.tabs.setCurrentWidget(self.monthly)
                self.statusBar().showMessage("月度统计已按调整后数据刷新")

    def schedule_adjustment_refresh(self, *_args):
        self.adjustment_refresh_timer.start()

    def render_adjustment_only(self, configs: list[AdjustmentConfig] | None = None):
        all_transactions = self.cached_transactions
        if not all_transactions:
            all_transactions, _duplicate_issues = dedupe_transactions([tx for result in self.results for tx in result.transactions])
            self.cached_transactions = all_transactions
        self.adjustment_result = apply_adjustments(all_transactions, configs or self.adjustment_configs())
        self.update_monthly_tab_label(self.adjustment_result.enabled)
        self.update_adjustment_preview()
        self._set_table(
            self.monthly,
            monthly_headers(self.adjustment_result.enabled, include_balances=False),
            build_monthly_display_rows(all_transactions, self.adjustment_result, include_balances=False),
        )
        if self.adjustment_result.enabled and self.adjustment_result.warnings:
            self.statusBar().showMessage("；".join(self.adjustment_result.warnings))

    def update_monthly_tab_label(self, adjusted: bool):
        self.tabs.setTabText(self.monthly_tab_index, "调整月度统计" if adjusted else "月度统计")

    def update_adjustment_preview(self):
        if not self.adjustment_result.enabled:
            self.adjust_net_value.setText("0.00")
            self.adjust_check_value.setText("未启用")
            self.adjust_preview_label.setText("原始统计保持不变")
            self.update_profit_preview()
            return
        total_row = next((row for row in self.adjustment_result.rows if row.month == "总计"), None)
        if total_row is None:
            self.adjust_net_value.setText("0.00")
            self.adjust_check_value.setText("需复核")
            self.adjust_preview_label.setText("暂无可调整月份")
            self.update_profit_preview()
            return
        check = balance_check(total_row, self.adjustment_result)
        self.adjust_net_value.setText(f"{money_wan(total_row.adjusted_net)}")
        self.adjust_check_value.setText(check)
        self.adjust_preview_label.setText(
            f"收入调整 {money_wan(total_row.income_adjustment)} 万元  "
            f"支出调整 {money_wan(total_row.expense_adjustment)} 万元"
        )
        self.update_profit_preview()

    def update_profit_preview(self):
        transactions = self.current_transactions()
        income_total, month_count = self.current_income_total_and_month_count(transactions)
        monthly_income = (
            income_total / Decimal(FIXED_PROOF_MONTHS) if month_count else Decimal("0.00")
        ).quantize(Decimal("0.01"))
        share_ratio = self.safe_percent(self.share_ratio.text() or "100")
        profit_rate = self.safe_percent(self.profit_rate.text())
        declared_income = self.safe_amount_wan(self.declared_month_income.text()) * Decimal("10000")
        generated_income = (
            monthly_income * share_ratio / Decimal("100") * profit_rate / Decimal("100")
        ).quantize(Decimal("0.01"))
        formula_hint = self.profit_formula_hint(month_count, share_ratio, profit_rate)

        self.profit_base_value.setText(money_wan(monthly_income))
        self.profit_generated_value.setText(money_wan(generated_income))
        if not transactions or month_count == 0:
            self.set_profit_check("待处理", "neutral")
            self.set_profit_value_tone("neutral")
            self.profit_hint_label.setText("处理流水后显示利润率校验")
            return
        if declared_income <= Decimal("0.00"):
            self.set_profit_check("请输入", "neutral")
            self.set_profit_value_tone("neutral")
            self.profit_hint_label.setText(formula_hint)
            return
        diff = (generated_income - declared_income).quantize(Decimal("0.01"))
        if diff >= Decimal("0.00"):
            self.set_profit_check("通过", "ok")
            self.set_profit_value_tone("ok")
            self.profit_hint_label.setText(f"{formula_hint}，余量 {money_wan(diff)}")
        else:
            self.set_profit_check("不足", "warning")
            self.set_profit_value_tone("warning")
            adjustment_hint = self.required_income_adjustment_hint(abs(diff), 6, share_ratio, profit_rate)
            self.profit_hint_label.setText(f"{formula_hint}，低于 {money_wan(abs(diff))}{adjustment_hint}")

    def set_profit_check(self, text: str, tone: str):
        self.profit_check_value.setText(text)
        self.profit_check_value.setProperty("tone", tone)
        self.profit_check_value.style().unpolish(self.profit_check_value)
        self.profit_check_value.style().polish(self.profit_check_value)

    def set_profit_value_tone(self, tone: str):
        self.profit_generated_value.setProperty("tone", tone)
        self.profit_generated_value.style().unpolish(self.profit_generated_value)
        self.profit_generated_value.style().polish(self.profit_generated_value)

    def profit_formula_hint(self, month_count: int, share_ratio: Decimal, profit_rate: Decimal) -> str:
        share_text = self.format_percent(share_ratio)
        profit_text = self.format_percent(profit_rate)
        if share_ratio == Decimal("100"):
            return f"识别到 {month_count} 个月；佐证按流水收入合计 ÷ 6 × {profit_text} 计算"
        return f"识别到 {month_count} 个月；佐证按流水收入合计 ÷ 6 × 占股{share_text} × {profit_text} 计算"

    def required_income_adjustment_hint(
        self,
        shortage: Decimal,
        month_count: int,
        share_ratio: Decimal,
        profit_rate: Decimal,
    ) -> str:
        factor = share_ratio / Decimal("100") * profit_rate / Decimal("100")
        if shortage <= Decimal("0.00") or month_count <= 0 or factor <= Decimal("0.00"):
            return ""
        monthly_income_needed = (shortage / factor).quantize(Decimal("0.01"))
        total_adjustment_needed = (monthly_income_needed * Decimal(month_count)).quantize(Decimal("0.01"))
        return (
            f"，需增加月均流水收入 {money_wan(monthly_income_needed)} 万元"
            f"，调整金额约 {money_wan(total_adjustment_needed)} 万元"
        )

    def format_percent(self, value: Decimal) -> str:
        normalized = value.quantize(Decimal("0.01")).normalize()
        text = format(normalized, "f")
        return f"{text}%"

    def current_transactions(self) -> list:
        if self.cached_transactions:
            return self.cached_transactions
        transactions = [tx for result in self.results for tx in result.transactions]
        deduped, _duplicate_issues = dedupe_transactions(transactions)
        self.cached_transactions = deduped
        return deduped

    def current_income_total_and_month_count(self, transactions: list) -> tuple[Decimal, int]:
        if not transactions:
            return Decimal("0.00"), 0
        if self.adjustment_result.enabled:
            total_row = next((row for row in self.adjustment_result.rows if row.month == "总计"), None)
            month_count = len([row for row in self.adjustment_result.rows if row.month != "总计"])
            if total_row is not None and month_count > 0:
                return total_row.adjusted_income_sum.quantize(Decimal("0.01")), month_count
        month_pairs = monthly_summaries(transactions)
        month_count = len(month_pairs)
        if month_count == 0:
            return Decimal("0.00"), 0
        total = summarize(transactions, "全部文件")
        return total.income_sum.quantize(Decimal("0.01")), month_count

    def render_empty(self):
        self.update_monthly_tab_label(False)
        self.adjustment_result = AdjustmentResult()
        self.cached_transactions = []
        self.update_adjustment_preview()
        for key, value in {
            "files": "0",
            "flows": "0",
            "income": "0 / 0.00",
            "expense": "0 / 0.00",
            "issues": "0",
            "confidence": "-",
        }.items():
            self.set_metric(key, value, "neutral" if key == "confidence" else None)
        self.summary_label.setText("选择 PDF/Excel 文件或文件夹后开始处理；默认输出近半年，可手动修改日期范围。")
        self._set_table(self.overview, ["文件", "状态"], [])
        self._set_table(self.monthly, ["月份", "收入(万元)", "支出(万元)"], [])
        self._set_table(self.details, ["时间", "收入", "支出", "余额"], [])
        self._set_table(self.issue_table, ["级别", "来源", "时间", "提示"], [])

    def render_selected(self):
        self.update_monthly_tab_label(False)
        self.adjustment_result = AdjustmentResult()
        self.cached_transactions = []
        self.update_adjustment_preview()
        self.set_metric("files", str(len(self.paths)))
        self.set_metric("flows", "待处理")
        self.set_metric("income", "待处理")
        self.set_metric("expense", "待处理")
        self.set_metric("issues", "待处理")
        self.set_metric("confidence", "待处理", "neutral")
        rows = [[str(path), "待处理"] for path in self.paths]
        self._set_table(self.overview, ["文件", "状态"], rows)
        self.summary_label.setText(f"已选择 {len(self.paths)} 个文件，点击开始处理。")

    def render_results(self):
        all_transactions, duplicate_issues = dedupe_transactions([tx for result in self.results for tx in result.transactions])
        self.cached_transactions = all_transactions
        shown_issues = self.issues + duplicate_issues
        total = summarize(all_transactions, "全部文件")
        type_counts = flow_type_counts(self.results)
        self.set_metric("files", str(len(self.results)))
        self.set_metric("flows", str(total.count))
        self.set_metric("income", f"{total.income_count} / {money(total.income_sum)}")
        self.set_metric("expense", f"{total.expense_count} / {money(total.expense_sum)}")
        self.set_metric("issues", str(len(shown_issues)))
        confidence = calculate_confidence(self.results, shown_issues, self.selected_date_range())
        self.set_metric("confidence", confidence.level, confidence.tone)
        self.summary_label.setText(
            f"文件 {len(self.results)} 个，流水 {total.count} 笔，收入 {total.income_count} 笔/{money(total.income_sum)}，"
            f"支出 {total.expense_count} 笔/{money(total.expense_sum)}，净额 {money(total.net)}，"
            f"异常提示 {len(shown_issues)} 条，可信度 {confidence.level}。类型：{type_counts}。"
        )

        self.adjustment_result = apply_adjustments(all_transactions, self.adjustment_configs())
        self.update_monthly_tab_label(self.adjustment_result.enabled)
        self.update_adjustment_preview()
        self._set_table(
            self.monthly,
            monthly_headers(self.adjustment_result.enabled, include_balances=False),
            build_monthly_display_rows(all_transactions, self.adjustment_result, include_balances=False),
        )
        if self.adjustment_result.enabled:
            self.summary_label.setText(self.summary_label.text() + " 已启用调整，月度统计显示调整后数据。")
            if self.adjustment_result.warnings:
                self.statusBar().showMessage("；".join(self.adjustment_result.warnings))

        overview_rows = []
        for result in self.results:
            s = result.summary
            overview_rows.append(
                [
                    result.path.name,
                    result_flow_type(result),
                    result_bank_label(result),
                    s.count,
                    s.income_count,
                    money(s.income_sum),
                    s.expense_count,
                    money(s.expense_sum),
                    money(s.net),
                    money(s.opening_balance),
                    money(s.closing_balance),
                    result.status,
                    result.message,
                ]
            )
        self._set_table(
            self.overview,
            ["文件", "流水类型", "银行", "流水笔数", "收入笔数", "收入", "支出笔数", "支出", "净额", "期初余额", "期末余额", "状态", "说明"],
            overview_rows,
        )

        detail_rows = []
        for tx in sort_transactions(all_transactions):
            detail_rows.append(
                [
                    tx.transaction_time.strftime("%Y-%m-%d %H:%M:%S"),
                    getattr(tx, "source_file", ""),
                    transaction_flow_type(tx),
                    getattr(tx, "bank_label", tx.bank),
                    money(tx.income),
                    money(tx.expense),
                    money(tx.balance),
                    tx.status,
                    "; ".join(tx.issues),
                    tx.raw_amount,
                    tx.raw_balance,
                ]
            )
        self._set_table(self.details, ["时间", "文件", "流水类型", "银行", "收入", "支出", "余额", "状态", "提示", "原始金额", "原始余额"], detail_rows)

        issue_rows = [[issue.level, issue.source, issue.time, issue.message, issue.raw_amount, issue.raw_balance] for issue in shown_issues]
        self._set_table(self.issue_table, ["级别", "来源", "时间", "提示", "原始金额", "原始余额"], issue_rows)

    def _set_table(self, table: QTableWidget, headers: list[str], rows: list[list]):
        table.setSortingEnabled(False)
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            row = list(row)
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            is_warning = any(str(value) == "需复核" for value in row)
            for col_index, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if is_warning:
                    item.setBackground(Qt.GlobalColor.yellow)
                table.setItem(row_index, col_index, item)
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        table.setSortingEnabled(True)

    def export_excel(self):
        if not self.results:
            QMessageBox.information(self, "提示", "请先处理流水后再导出。")
            return
        file_name, _ = QFileDialog.getSaveFileName(self, "导出 Excel", default_export_path(self.results), "Excel 文件 (*.xlsx)")
        if not file_name:
            return
        path = Path(file_name)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)
        write_workbook(path, self.results, self.issues, self.adjustment_configs())
        QMessageBox.information(self, "完成", f"已导出: {path}")

    def export_income_proof_json(self):
        if not self.results:
            opened = open_income_proof_form()
            if not opened:
                QMessageBox.warning(self, "未打开填表", "未找到收入佐证程序，请确认工具包目录完整。")
            return
        default_path = Path(default_export_path(self.results)).with_suffix(".income_proof.json")
        path = default_path
        path.parent.mkdir(parents=True, exist_ok=True)
        write_income_proof_input(path, self.results, adjustment_configs=self.adjustment_configs())
        opened = open_income_proof_form(path)
        if not opened:
            QMessageBox.warning(self, "未打开填表", f"已导出: {path}\n未找到收入佐证程序，请确认工具包目录完整。")

    def export_salary_income_proof_json(self):
        default_path = Path(default_export_path(self.results)).with_suffix(".salary_income_proof.json")
        path = default_path
        path.parent.mkdir(parents=True, exist_ok=True)
        write_salary_income_proof_input(path, self.results, adjustment_configs=self.adjustment_configs())
        opened = open_income_proof_form(path)
        if not opened:
            QMessageBox.warning(self, "未打开填表", f"已导出: {path}\n未找到收入佐证程序，请确认工具包目录完整。")

    def adjustment_configs(self) -> list[AdjustmentConfig]:
        start_month, end_month = self.adjustment_month_range()
        return [
            AdjustmentConfig(
                enabled=self.income_adjust.isChecked(),
                amount_wan=self.safe_amount_wan(self.adjust_amount.text()),
                start_month=start_month,
                end_month=end_month,
                balanced=False,
                label="收入调整（微信）",
                randomized=not self.random_adjust.isChecked(),
            ),
            AdjustmentConfig(
                enabled=self.balance_adjust.isChecked(),
                amount_wan=self.safe_amount_wan(self.adjust_amount.text()),
                start_month=start_month,
                end_month=end_month,
                balanced=True,
                label="收支平衡调整（个/公）",
                randomized=not self.random_adjust.isChecked(),
            ),
        ]

    def adjustment_month_range(self) -> tuple[str, str]:
        if not self.date_filter.isChecked():
            return "", ""
        return self.start_date.date().toString("yyyy-MM"), self.end_date.date().toString("yyyy-MM")

    def safe_amount_wan(self, text: str) -> Decimal:
        try:
            return parse_amount_wan(text)
        except Exception:
            return Decimal("0.00")

    def safe_percent(self, text: str) -> Decimal:
        try:
            value = Decimal((text or "").replace("%", "").strip() or "0")
        except Exception:
            return Decimal("0.00")
        return max(value, Decimal("0.00"))


def write_sheet(ws, headers: list[str], rows: list[list]):
    from openpyxl.styles import Font, PatternFill

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E9EEF5", end_color="E9EEF5", fill_type="solid")
    for row in rows:
        ws.append(row)
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(max(width, 10), 42)


def default_export_filename(results: list[FileResult]) -> str:
    names: list[str] = []
    for result in results:
        name = safe_filename_part(result.account_name)
        if name and name not in names:
            names.append(name)
    if not names:
        return "银行流水解析结果.xlsx"
    return f"银行流水解析结果_{'_'.join(names[:3])}.xlsx"


def default_export_path(results: list[FileResult]) -> str:
    filename = default_export_filename(results)
    export_dir = default_export_dir()
    if export_dir:
        return str(export_dir / filename)
    return filename


def default_export_dir() -> Path | None:
    return runtime_dir() / "银行流水解析结果"


def safe_filename_part(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "", value or "").strip()[:30]


def extract_account_name(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        return extract_pdf_account_name(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return extract_excel_account_name(path)
    return ""


def extract_account_no(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        return extract_pdf_account_no(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return extract_excel_account_no(path)
    return ""


def extract_pdf_account_name(path: Path) -> str:
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            if not pdf.pages:
                return ""
            return parse_account_name(pdf.pages[0].extract_text() or "")
    except Exception:
        return ""


def extract_pdf_account_no(path: Path) -> str:
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            if not pdf.pages:
                return ""
            chunks = []
            for page in pdf.pages[:2]:
                chunks.append(page.extract_text() or "")
            return parse_account_no("\n".join(chunks))
    except Exception:
        return ""


def extract_excel_account_name(path: Path) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            chunks = []
            for row in worksheet.iter_rows(min_row=1, max_row=30, values_only=True):
                chunks.append(" ".join(str(cell or "") for cell in row))
            return parse_account_name("\n".join(chunks))
        finally:
            workbook.close()
    except Exception:
        return ""


def extract_excel_account_no(path: Path) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            chunks = []
            for row in worksheet.iter_rows(min_row=1, max_row=30, values_only=True):
                chunks.append(" ".join(str(cell or "") for cell in row))
            return parse_account_no("\n".join(chunks))
        finally:
            workbook.close()
    except Exception:
        return ""


def parse_account_name(text: str) -> str:
    for pattern in ACCOUNT_NAME_PATTERNS:
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
    return ""


def normalize_account_no(value: str) -> str:
    return re.sub(r"[^0-9*＊]", "", value or "").replace("＊", "*")


def valid_account_no(value: str) -> bool:
    if not 8 <= len(value) <= 22:
        return False
    if re.fullmatch(r"20\d{6}", value):
        return False
    return bool(re.search(r"\d", value))


def parse_account_no(text: str) -> str:
    header_lines = []
    for raw_line in (text or "").splitlines()[:80]:
        line = raw_line.strip()
        if not line or "对方账号" in line or "对手账号" in line:
            continue
        header_lines.append(line)
    header_text = "\n".join(header_lines)
    for pattern in ACCOUNT_NO_PATTERNS:
        match = re.search(pattern, header_text, re.I)
        if match:
            account_no = normalize_account_no(match.group(1))
            if valid_account_no(account_no):
                return account_no
    return ""


def write_workbook(path: Path, results: list[FileResult], issues: list[Issue], adjustment_configs: list[AdjustmentConfig] | None = None):
    from openpyxl import Workbook

    wb = Workbook()
    monthly = wb.active
    monthly.title = "原始月度统计"
    overview = wb.create_sheet("汇总")
    overview_rows = []
    all_transactions = []
    for result in results:
        s = result.summary
        overview_rows.append(
            [
                result.path.name,
                result_flow_type(result),
                result.bank_label,
                s.count,
                s.income_count,
                float(s.income_sum),
                s.expense_count,
                float(s.expense_sum),
                float(s.net),
                float(s.opening_balance) if s.opening_balance is not None else None,
                float(s.closing_balance) if s.closing_balance is not None else None,
                result.status,
                result.message,
            ]
        )
        all_transactions.extend(result.transactions)
    write_sheet(overview, ["文件", "流水类型", "银行", "流水笔数", "收入笔数", "收入", "支出笔数", "支出", "净额", "期初余额", "期末余额", "状态", "说明"], overview_rows)

    all_transactions, duplicate_issues = dedupe_transactions(all_transactions)
    shown_issues = issues + duplicate_issues
    monthly_rows = build_monthly_display_rows(all_transactions, AdjustmentResult(), for_excel=True)
    write_sheet(monthly, monthly_headers(False), monthly_rows)

    adjustment_result = apply_adjustments(all_transactions, adjustment_configs or [])
    adjusted = wb.create_sheet("调整后月度统计")
    write_sheet(adjusted, monthly_headers(True), [["全部"] + row for row in build_adjusted_rows(adjustment_result, for_excel=True)])

    parameters = wb.create_sheet("调整参数")
    write_sheet(parameters, ["参数", "值"], adjustment_result.parameters)

    instruction_rows = [
        ["说明", "调整功能属于识别后的测算层，不修改原始明细。"],
        ["收入调整（微信）", "只增加收入，适合微信或需要测算收入补充的流水。"],
        ["收支平衡调整（个/公）", "收入和支出同步增加，期末余额保持原值。"],
        ["月份范围", "只调整已有流水的月份，不生成没有流水的月份。"],
        ["余额判断", "调整后余额 >= 0 为正常，否则需复核。"],
        ["微信其他", "微信“其他”在原始统计中默认排除。"],
    ]
    instructions = wb.create_sheet("调整说明")
    write_sheet(instructions, ["项目", "说明"], instruction_rows)

    details = wb.create_sheet("原始明细")
    detail_rows = []
    for tx in sort_transactions(all_transactions):
        detail_rows.append(
            [
                tx.transaction_time.strftime("%Y-%m-%d %H:%M:%S"),
                getattr(tx, "source_file", ""),
                transaction_flow_type(tx),
                getattr(tx, "bank_label", tx.bank),
                float(tx.income),
                float(tx.expense),
                float(tx.balance) if tx.balance is not None else None,
                tx.status,
                "; ".join(tx.issues),
                tx.raw_amount,
                tx.raw_balance,
            ]
        )
    write_sheet(details, ["时间", "文件", "流水类型", "银行", "收入", "支出", "余额", "状态", "提示", "原始金额", "原始余额"], detail_rows)

    issue_sheet = wb.create_sheet("异常提示")
    issue_rows = [[issue.level, issue.source, issue.time, issue.message, issue.raw_amount, issue.raw_balance] for issue in shown_issues]
    write_sheet(issue_sheet, ["级别", "来源", "时间", "提示", "原始金额", "原始余额"], issue_rows)

    wb.save(path)


def dedupe_transactions(transactions: list) -> tuple[list, list[Issue]]:
    unique = []
    seen: dict[tuple, object] = {}
    duplicate_groups: dict[tuple[str, str], dict[str, object]] = {}

    for tx in sorted(transactions, key=lambda item: (item.transaction_time, getattr(item, "source_file", ""), item.row_no)):
        key = getattr(tx, "merge_key", None)
        if key:
            signature = (tx.bank, key)
        else:
            signature = (
                tx.bank,
                tx.transaction_time,
                tx.income,
                tx.expense,
                tx.balance,
                tx.raw_amount,
                tx.raw_balance,
            )

        if signature in seen and (key or getattr(seen[signature], "source_file", "") != getattr(tx, "source_file", "")):
            first = seen[signature]
            group_key = (getattr(tx, "source_file", ""), getattr(first, "source_file", ""))
            group = duplicate_groups.setdefault(
                group_key,
                {
                    "count": 0,
                    "time": tx.transaction_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "raw_amount": tx.raw_amount,
                    "raw_balance": tx.raw_balance,
                },
            )
            group["count"] = int(group["count"]) + 1
            continue

        if signature not in seen:
            seen[signature] = tx
        unique.append(tx)

    issues = [
        Issue(
            "需复核",
            source,
            str(group["time"]),
            f"疑似重复流水 {group['count']} 笔，已在合并明细中去重；首次来源: {first_source}",
            str(group["raw_amount"]),
            str(group["raw_balance"]),
        )
        for (source, first_source), group in duplicate_groups.items()
    ]
    return unique, issues


def result_flow_type(result: FileResult) -> str:
    return infer_flow_type(result.bank_id, result.account_name, result.transactions)


def result_bank_label(result: FileResult) -> str:
    if result.bank_label:
        return result.bank_label
    from bankflow_v2.auto_detect import BANK_LABELS

    return BANK_LABELS.get(result.bank_id, "")


def transaction_flow_type(tx) -> str:
    value = getattr(tx, "flow_type", "")
    return value or "个人"


def infer_excel_bank_label(bank_label: str, transactions: list) -> str:
    if bank_label != "Excel导入":
        return bank_label
    labels = {
        getattr(tx, "bank", "")
        for tx in transactions
        if getattr(tx, "bank", "") and getattr(tx, "bank", "") != "Excel导入"
    }
    if len(labels) == 1:
        return next(iter(labels))
    return bank_label


def infer_flow_type(bank_id: str, account_name: str = "", transactions: list | None = None) -> str:
    detected_flow_type = income_flow_type(bank_id)
    if detected_flow_type == "个人" and looks_corporate_account_name(account_name):
        return "对公"
    if bank_id == "excel":
        labels = " ".join(
            str(value)
            for tx in (transactions or [])
            for value in (getattr(tx, "bank", ""), getattr(tx, "bank_label", ""))
            if value
        )
        if "微信" in labels:
            return "微信"
        if "对公" in labels:
            return "对公"
    return detected_flow_type


def flow_type_counts(results: list[FileResult]) -> str:
    counts = {"个人": 0, "微信": 0, "对公": 0}
    for result in results:
        counts[result_flow_type(result)] = counts.get(result_flow_type(result), 0) + 1
    return "，".join(f"{key}{value}份" for key, value in counts.items() if value) or "无"


def calculate_confidence(results: list[FileResult], issues: list[Issue], date_range: tuple[datetime | None, datetime | None]) -> ConfidenceInfo:
    if not results:
        return ConfidenceInfo(0, "-", "neutral", "暂无结果")

    scored_results = [result for result in results if not is_date_range_empty_result(result)]
    scored_issues = [issue for issue in issues if not is_date_range_empty_issue(issue)]
    if not scored_results:
        return ConfidenceInfo(0, "-", "neutral", DATE_RANGE_EMPTY_MESSAGE)
    if not any(int(getattr(result.summary, "count", 0)) > 0 for result in scored_results):
        return ConfidenceInfo(0, "未识别", "confidenceLow", "未识别到流水")

    weighted_score = 0
    total_weight = 0
    reasons: list[str] = []
    for result in scored_results:
        score, reason = result_confidence_score(result, date_range)
        weight = max(int(getattr(result.summary, "count", 0)), 1)
        weighted_score += score * weight
        total_weight += weight
        reasons.append(reason)

    score = round(weighted_score / total_weight) if total_weight else 0
    review_count = sum(1 for issue in scored_issues if issue.level == "需复核")
    low_risk_count = sum(1 for issue in scored_issues if issue.level != "需复核")
    score -= min(18, review_count * 4)
    score -= min(6, low_risk_count)
    score = max(0, min(100, score))

    if is_extremely_high_confidence(scored_results, scored_issues):
        level = "极高"
        tone = "confidenceVeryHigh"
    elif score >= 90:
        level = "高"
        tone = "confidenceHigh"
    elif score >= 70:
        level = "中"
        tone = "confidenceMedium"
    else:
        level = "低"
        tone = "confidenceLow"
    return ConfidenceInfo(score, level, tone, "；".join(reasons))


def result_confidence_score(result: FileResult, date_range: tuple[datetime | None, datetime | None]) -> tuple[int, str]:
    summary = result.summary
    count = int(getattr(summary, "count", 0))
    if count <= 0:
        return 0, "未识别到流水"

    used_generic = is_generic_confidence_result(result)
    if used_generic:
        score = 68
        reason = "通用识别"
    elif result.bank_id == "excel":
        score = 86
        reason = "Excel导入"
    else:
        score = min(94, max(78, int(result.bank_confidence or 80)))
        reason = "专用解析"

    review_issues = [issue for issue in getattr(summary, "issues", []) if issue.level == "需复核"]
    low_risk_issues = [issue for issue in getattr(summary, "issues", []) if issue.level != "需复核"]
    score -= min(22, len(review_issues) * 6)
    score -= min(8, len(low_risk_issues) * 2)

    if not confidence_needs_balance(result):
        score += 6
    elif getattr(summary, "opening_balance", None) is not None and getattr(summary, "closing_balance", None) is not None:
        has_balance_break = has_review_balance_break(summary)
        if summary_balance_closed(summary):
            score += 26 if not has_balance_break else 10
        else:
            score -= 24
    else:
        score -= 6

    if last_two_months_have_flows(result.transactions, date_range):
        score += 8
        if last_two_months_closed_without_review(result.transactions, date_range):
            score += 14

    if result.message and result.status == "需复核" and result.message != DATE_RANGE_EMPTY_MESSAGE:
        score -= 8

    return max(0, min(100, score)), reason


def is_extremely_high_confidence(results: list[FileResult], issues: list[Issue]) -> bool:
    if issues:
        return False
    for result in results:
        if int(getattr(result.summary, "count", 0)) <= 0:
            return False
        if result.bank_id == "excel":
            return False
        if is_generic_confidence_result(result) and not summary_balance_closed(result.summary):
            return False
        if getattr(result.summary, "issues", []):
            return False
        if result.message and result.message != DATE_RANGE_EMPTY_MESSAGE:
            return False
    return True


def only_total_balance_closure_issues(results: list[FileResult], issues: list[Issue]) -> bool:
    if not issues:
        return False
    if any(int(getattr(result.summary, "count", 0)) <= 0 for result in results):
        return False
    return all(is_total_balance_closure_issue(issue) for issue in issues)


def is_total_balance_closure_issue(issue: Issue) -> bool:
    return issue.level == "需复核" and str(issue.message).startswith("收支余额不闭合")


def is_generic_confidence_result(result: FileResult) -> bool:
    return (
        result.status == "通用识别"
        or "通用识别" in (result.bank_label or "")
        or result.bank_id in ("generic_pdf", "cmbc", "cib")
    )


def summary_balance_closed(summary) -> bool:
    opening = getattr(summary, "opening_balance", None)
    closing = getattr(summary, "closing_balance", None)
    if opening is None or closing is None:
        return False
    expected_change = (closing - opening).quantize(Decimal("0.01"))
    return getattr(summary, "net", Decimal("0.00")).quantize(Decimal("0.01")) == expected_change


def has_review_balance_break(summary) -> bool:
    return any(
        "余额不连续" in issue.message and issue.level == "需复核"
        for issue in getattr(summary, "issues", [])
    )


def confidence_needs_balance(result: FileResult) -> bool:
    if result.bank_id == "wechat":
        return False
    if result.transactions and all(getattr(tx, "balance_optional", False) for tx in result.transactions):
        return False
    return True


def is_date_range_empty_result(result: FileResult) -> bool:
    return int(getattr(result.summary, "count", 0)) == 0 and result.message == DATE_RANGE_EMPTY_MESSAGE


def is_date_range_empty_issue(issue: Issue) -> bool:
    return issue.message == DATE_RANGE_EMPTY_MESSAGE


def last_two_months_have_flows(transactions: list, date_range: tuple[datetime | None, datetime | None]) -> bool:
    target_months = last_two_month_keys(date_range)
    if len(target_months) < 2 and transactions:
        latest = max(tx.transaction_time for tx in transactions)
        month_start = datetime(latest.year, latest.month, 1)
        target_months = [add_months(month_start, -1).strftime("%Y-%m"), month_start.strftime("%Y-%m")]
    if len(target_months) < 2:
        return False
    seen = {tx.transaction_time.strftime("%Y-%m") for tx in transactions}
    return all(month in seen for month in target_months)


def last_two_months_closed_without_review(transactions: list, date_range: tuple[datetime | None, datetime | None]) -> bool:
    target_months = last_two_month_keys(date_range)
    if len(target_months) < 2 and transactions:
        latest = max(tx.transaction_time for tx in transactions)
        month_start = datetime(latest.year, latest.month, 1)
        target_months = [add_months(month_start, -1).strftime("%Y-%m"), month_start.strftime("%Y-%m")]
    if len(target_months) < 2:
        return False

    for month in target_months:
        month_transactions = [tx for tx in transactions if tx.transaction_time.strftime("%Y-%m") == month]
        if not month_transactions:
            return False
        month_summary = summarize(month_transactions, month)
        if any(issue.level == "需复核" for issue in month_summary.issues):
            return False
        if confidence_needs_balance_for_transactions(month_transactions) and not summary_balance_closed(month_summary):
            return False
    return True


def confidence_needs_balance_for_transactions(transactions: list) -> bool:
    if not transactions:
        return True
    if all(getattr(tx, "bank", "") == "微信流水" for tx in transactions):
        return False
    if all(getattr(tx, "balance_optional", False) for tx in transactions):
        return False
    return True


def last_two_month_keys(date_range: tuple[datetime | None, datetime | None]) -> list[str]:
    _start, end = date_range
    if end is None:
        if not date_range[0]:
            return []
        end = max(date_range[0], datetime.now())
    month_start = datetime(end.year, end.month, 1)
    previous = add_months(month_start, -1)
    return [previous.strftime("%Y-%m"), month_start.strftime("%Y-%m")]


def add_months(value: datetime, months: int) -> datetime:
    month_index = value.year * 12 + value.month - 1 + months
    year = month_index // 12
    month = month_index % 12 + 1
    return datetime(year, month, 1)


def grouped_transactions(transactions: list) -> list[tuple[str, list]]:
    order = ["个人", "微信", "对公"]
    groups = []
    for flow_type in order:
        items = [tx for tx in transactions if transaction_flow_type(tx) == flow_type]
        if items:
            groups.append((flow_type, items))
    return groups


def build_monthly_rows(transactions: list, for_excel: bool = False, balance_wechat: bool = False) -> list[list]:
    rows = []
    month_pairs = monthly_summaries(transactions)
    if balance_wechat:
        month_pairs = list(balance_wechat_summaries(month_pairs).items())
    for month, s in month_pairs:
        rows.append(_summary_row("", month, s, for_excel))

    if rows:
        if balance_wechat:
            total = sum_summaries([s for _month, s in month_pairs])
        else:
            total = summarize(transactions, "总计")
        rows.append(_summary_row("", "总计", total, for_excel))
        rows.append(_monthly_average_row(total, len(rows) - 1, for_excel, adjusted=False))
    return rows


def monthly_headers(adjusted: bool = False, include_balances: bool = True) -> list[str]:
    headers = [
        "流水类型",
        "月份",
        "收入笔数",
        "收入(万元)",
        "支出笔数",
        "支出(万元)",
        "净额(万元)",
        "流水笔数",
    ]
    if include_balances:
        headers[7:7] = ["期初余额(万元)", "期末余额(万元)"]
    if adjusted:
        headers.extend([
            "收入调整(万元)",
            "支出调整(万元)",
            "调整后余额变动(万元)",
            "调整后收支差额(万元)",
            "平衡校验",
            "说明",
        ])
    return headers


def build_monthly_display_rows(
    transactions: list,
    adjustment: AdjustmentResult,
    for_excel: bool = False,
    include_balances: bool = True,
) -> list[list]:
    if not adjustment.enabled:
        rows = build_grouped_monthly_display_rows(transactions, for_excel)
    else:
        rows = build_grouped_monthly_display_rows(transactions, for_excel, include_combined=False)
        rows.extend([["全部"] + row for row in build_adjusted_rows(adjustment, for_excel)])
    if not include_balances:
        rows = [without_monthly_balance_columns(row) for row in rows]
    return rows


def build_grouped_monthly_display_rows(transactions: list, for_excel: bool = False, include_combined: bool = True) -> list[list]:
    rows = []
    groups = grouped_transactions(transactions)
    for flow_type, items in groups:
        group_rows = build_monthly_rows(items, for_excel, balance_wechat=flow_type == "微信")
        for row in group_rows:
            row[0] = flow_type
        rows.extend(group_rows)
    if include_combined and len(groups) > 1:
        total_rows = build_combined_monthly_rows(groups, for_excel)
        for row in total_rows:
            row[0] = "全部"
        rows.extend(total_rows)
    return rows


def without_monthly_balance_columns(row: list) -> list:
    return [value for index, value in enumerate(row) if index not in (7, 8)]


def build_combined_monthly_rows(groups: list[tuple[str, list]], for_excel: bool = False) -> list[list]:
    summaries_by_month: dict[str, list[Summary]] = {}
    for flow_type, items in groups:
        month_pairs = monthly_summaries(items)
        if flow_type == "微信":
            month_pairs = list(balance_wechat_summaries(month_pairs).items())
        for month, summary in month_pairs:
            summaries_by_month.setdefault(month, []).append(summary)

    rows = []
    combined_pairs = [(month, sum_summaries(summaries)) for month, summaries in sorted(summaries_by_month.items())]
    for month, summary in combined_pairs:
        rows.append(_summary_row("", month, summary, for_excel))
    if rows:
        total = sum_summaries([summary for _month, summary in combined_pairs])
        rows.append(_summary_row("", "总计", total, for_excel))
        rows.append(_monthly_average_row(total, len(rows) - 1, for_excel, adjusted=False))
    return rows


def sum_summaries(summaries: list[Summary]) -> Summary:
    total = Summary()
    for summary in summaries:
        total.count += summary.count
        total.income_count += summary.income_count
        total.income_sum += summary.income_sum
        total.expense_count += summary.expense_count
        total.expense_sum += summary.expense_sum
    total.income_sum = total.income_sum.quantize(Decimal("0.01"))
    total.expense_sum = total.expense_sum.quantize(Decimal("0.01"))
    total.net = (total.income_sum - total.expense_sum).quantize(Decimal("0.01"))
    return total


def build_adjusted_rows(adjustment: AdjustmentResult, for_excel: bool = False) -> list[list]:
    rows = []
    for row in adjustment.rows:
        if row.month == "总计":
            continue
        values = [
            row.adjusted_income_sum,
            row.adjusted_expense_sum,
            row.adjusted_net,
            row.adjusted_opening_balance,
            row.adjusted_closing_balance,
            row.income_adjustment,
            row.expense_adjustment,
            balance_delta(row.adjusted_opening_balance, row.adjusted_closing_balance),
            row.adjusted_net,
        ]
        if for_excel:
            money_values = [float(to_wan(value)) if value is not None else None for value in values]
        else:
            money_values = [money_wan(value) for value in values]
        rows.append(
            [
                row.month,
                row.original_income_count,
                money_values[0],
                row.original_expense_count,
                money_values[1],
                money_values[2],
                money_values[3],
                money_values[4],
                row.original_count,
                money_values[5],
                money_values[6],
                money_values[7],
                money_values[8],
                balance_check(row, adjustment),
                row.note,
            ]
        )
    for warning in adjustment.warnings:
        rows.append(["提示", "", "", "", "", "", "", "", "", "", "", "", "", "", warning])
    total_row = next((row for row in adjustment.rows if row.month == "总计"), None)
    if total_row is not None:
        rows.append(_adjusted_row(total_row, for_excel, label="总计", adjustment=adjustment))
        month_count = max(len([row for row in adjustment.rows if row.month != "总计"]), 1)
        rows.append(_monthly_average_row(total_row, month_count, for_excel, adjusted=True))
    return rows


def _adjusted_row(row, for_excel: bool, label: str | None = None, adjustment: AdjustmentResult | None = None) -> list:
    values = [
        row.adjusted_income_sum,
        row.adjusted_expense_sum,
        row.adjusted_net,
        row.adjusted_opening_balance,
        row.adjusted_closing_balance,
        row.income_adjustment,
        row.expense_adjustment,
        balance_delta(row.adjusted_opening_balance, row.adjusted_closing_balance),
        row.adjusted_net,
    ]
    if for_excel:
        money_values = [float(to_wan(value)) if value is not None else None for value in values]
    else:
        money_values = [money_wan(value) for value in values]
    return [
        label or row.month,
        row.original_income_count,
        money_values[0],
        row.original_expense_count,
        money_values[1],
        money_values[2],
        money_values[3],
        money_values[4],
        row.original_count,
        money_values[5],
        money_values[6],
        money_values[7],
        money_values[8],
        balance_check(row, adjustment or AdjustmentResult(balanced=row.expense_adjustment != Decimal("0.00"))),
        row.note,
    ]


def _monthly_average_row(total, month_count: int, for_excel: bool, adjusted: bool) -> list:
    divisor = Decimal(FIXED_PROOF_MONTHS)
    income_average = (total.adjusted_income_sum if adjusted else total.income_sum) / divisor
    expense_average = (total.adjusted_expense_sum if adjusted else total.expense_sum) / divisor
    income_average = income_average.quantize(Decimal("0.01"))
    expense_average = expense_average.quantize(Decimal("0.01"))
    if for_excel:
        income_value = float(to_wan(income_average))
        expense_value = float(to_wan(expense_average))
    else:
        income_value = money_wan(income_average)
        expense_value = money_wan(expense_average)
    if adjusted:
        return ["月均(÷6)", "", income_value, "", expense_value, "", "", "", "", "", "", "", "", "", f"识别{month_count}个月，固定按6个月平均"]
    return ["", "月均(÷6)", "", income_value, "", expense_value, "", "", "", ""]


def balance_delta(opening: Decimal | None, closing: Decimal | None) -> Decimal | None:
    if opening is None or closing is None:
        return None
    return (closing - opening).quantize(Decimal("0.01"))


def balance_check(row, adjustment: AdjustmentResult) -> str:
    if not adjustment.balanced:
        return "模拟"
    delta = balance_delta(row.adjusted_opening_balance, row.adjusted_closing_balance)
    if delta is None:
        return "需复核"
    return "正常" if delta == row.adjusted_net else "需复核"

def build_salary_sheet(transactions: list, for_excel: bool = False) -> tuple[list[str], list[list]]:
    salary_transactions = [tx for tx in sort_transactions(transactions) if is_salary_transaction(tx)]
    if not salary_transactions:
        return [], []

    headers = salary_headers(salary_transactions)
    amount_col = salary_amount_col(headers)
    rows = []
    for tx in salary_transactions:
        fields = salary_fields(tx)
        rows.append((fields + [""] * len(headers))[:len(headers)])

    total = sum((tx.income for tx in salary_transactions), Decimal("0.00")).quantize(Decimal("0.01"))
    months = {tx.transaction_time.strftime("%Y-%m") for tx in salary_transactions}
    average = (total / Decimal(len(months))).quantize(Decimal("0.01")) if months else Decimal("0.00")
    rows.append([""] * len(headers))
    rows.append(salary_summary_row(headers, amount_col, "工资总额", total, for_excel))
    rows.append(salary_summary_row(headers, amount_col, "统计月份数", Decimal(len(months)), for_excel))
    rows.append(salary_summary_row(headers, amount_col, "月均", average, for_excel))
    return headers, rows


def salary_headers(transactions: list) -> list[str]:
    headers = next((list(getattr(tx, "raw_headers", []) or []) for tx in transactions if getattr(tx, "raw_headers", None)), [])
    max_len = max(len(salary_fields(tx)) for tx in transactions)
    if not headers:
        headers = [f"原始字段{index}" for index in range(1, max_len + 1)]
    if len(headers) < max_len:
        headers.extend(f"原始字段{index}" for index in range(len(headers) + 1, max_len + 1))
    return headers


def salary_fields(tx) -> list:
    fields = list(getattr(tx, "raw_fields", []) or [])
    if fields:
        return fields
    text = salary_match_text(tx)
    return text.split(" | ") if " | " in text else [text]


def salary_amount_col(headers: list[str]) -> int:
    preferred = ("交易金额", "收入金额", "贷方发生额", "贷方", "金额")
    for name in preferred:
        for index, header in enumerate(headers):
            if name in header:
                return index
    return min(3, len(headers) - 1)


def salary_summary_row(headers: list[str], amount_col: int, label: str, value: Decimal, for_excel: bool) -> list:
    row = [""] * len(headers)
    row[0] = label
    if label == "统计月份数":
        row[amount_col] = int(value)
    else:
        row[amount_col] = float(value) if for_excel else money(value)
    return row


def is_salary_transaction(tx) -> bool:
    if tx.income <= Decimal("0.00"):
        return False
    text = salary_match_text(tx)
    return any(keyword in text for keyword in SALARY_KEYWORDS)


def salary_match_text(tx) -> str:
    raw_fields = " | ".join(str(field) for field in getattr(tx, "raw_fields", []) or [])
    return f"{getattr(tx, 'raw_text', '')} | {raw_fields} | {getattr(tx, 'raw_amount', '')}".strip()


def build_combined_summary_rows(results: list[FileResult], transactions: list, for_excel: bool = False) -> list[list]:
    rows = []
    for result in results:
        rows.append(_combined_summary_row(
            "文件",
            result.path.name,
            result_bank_label(result),
            result.summary,
            result.status,
            result.message,
            for_excel,
        ))

    month_summaries: dict[str, list[Summary]] = {}
    for result in results:
        for month, summary in monthly_summaries(getattr(result, "transactions", []) or []):
            month_summaries.setdefault(month, []).append(summary)
    for month, summaries in sorted(month_summaries.items()):
        summary = sum_summaries(summaries)
        rows.append(_combined_summary_row("月份", month, "全部文件", summary, "", "", for_excel))

    if month_summaries:
        rows.append(_combined_summary_row("总计", "总计", "全部文件", sum_summaries([result.summary for result in results]), "", "", for_excel))
    return rows


def _combined_summary_row(
    row_type: str,
    name: str,
    scope: str,
    s,
    status: str,
    message: str,
    for_excel: bool,
) -> list:
    values = [
        s.income_sum,
        s.expense_sum,
        s.net,
        s.opening_balance,
        s.closing_balance,
    ]
    if for_excel:
        money_values = [float(to_wan(value)) if value is not None else None for value in values]
    else:
        money_values = [money_wan(value) for value in values]
    return [
        row_type,
        name,
        scope,
        s.count,
        s.income_count,
        money_values[0],
        s.expense_count,
        money_values[1],
        money_values[2],
        money_values[3],
        money_values[4],
        status,
        message,
    ]


def _summary_row(flow_type: str, label: str, s, for_excel: bool) -> list:
    if for_excel:
        return [
            flow_type,
            label,
            s.income_count,
            float(to_wan(s.income_sum)),
            s.expense_count,
            float(to_wan(s.expense_sum)),
            float(to_wan(s.net)),
            float(to_wan(s.opening_balance)) if s.opening_balance is not None else None,
            float(to_wan(s.closing_balance)) if s.closing_balance is not None else None,
            s.count,
        ]
    return [
        flow_type,
        label,
        s.income_count,
        money_wan(s.income_sum),
        s.expense_count,
        money_wan(s.expense_sum),
        money_wan(s.net),
        money_wan(s.opening_balance),
        money_wan(s.closing_balance),
        s.count,
    ]


def to_wan(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return (value / Decimal("10000")).quantize(Decimal("0.01"))


def money_wan(value: Decimal | None) -> str:
    converted = to_wan(value)
    return "" if converted is None else f"{converted:,.2f}"


def default_recent_month_range() -> tuple[QDate, QDate]:
    today = QDate.currentDate()
    month_start = QDate(today.year(), today.month(), 1)
    end_month = month_start if today.day() >= 15 else month_start.addMonths(-1)
    start = end_month.addMonths(-5)
    end = QDate(end_month.year(), end_month.month(), end_month.daysInMonth())
    return start, end


def apply_light_palette(app: QApplication) -> None:
    app.setStyle("Fusion")
    palette = QPalette()
    roles = QPalette.ColorRole
    groups = QPalette.ColorGroup

    palette.setColor(roles.Window, QColor("#f5f7fb"))
    palette.setColor(roles.WindowText, QColor("#162033"))
    palette.setColor(roles.Base, QColor("#ffffff"))
    palette.setColor(roles.AlternateBase, QColor("#f8fbff"))
    palette.setColor(roles.ToolTipBase, QColor("#ffffff"))
    palette.setColor(roles.ToolTipText, QColor("#162033"))
    palette.setColor(roles.Text, QColor("#162033"))
    palette.setColor(roles.Button, QColor("#ffffff"))
    palette.setColor(roles.ButtonText, QColor("#263243"))
    palette.setColor(roles.BrightText, QColor("#ffffff"))
    palette.setColor(roles.Highlight, QColor("#cfe5ff"))
    palette.setColor(roles.HighlightedText, QColor("#162033"))
    palette.setColor(groups.Disabled, roles.WindowText, QColor("#9aa6b2"))
    palette.setColor(groups.Disabled, roles.Text, QColor("#9aa6b2"))
    palette.setColor(groups.Disabled, roles.ButtonText, QColor("#9aa6b2"))
    palette.setColor(groups.Disabled, roles.Highlight, QColor("#e7edf5"))
    palette.setColor(groups.Disabled, roles.HighlightedText, QColor("#8c99a8"))
    app.setPalette(palette)


def open_income_proof_form(json_path: Path | None = None) -> bool:
    candidates = []
    source_launcher = Path(r"D:\report workflow\启动收入佐证填表.bat")
    if source_launcher.exists():
        candidates.append(source_launcher)
    if is_packaged_app():
        app_dir = runtime_dir()
        candidates.extend([
            app_dir.parent / "收入佐证" / "IncomeProofGUI.exe",
            app_dir / "IncomeProofGUI.exe",
        ])
    else:
        app_dir = runtime_dir()
        candidates.extend([
            app_dir.parent / "收入佐证" / "启动收入佐证填表.bat",
            app_dir.parent / "收入佐证" / "IncomeProofGUI.exe",
        ])
    launcher = next((path for path in candidates if path.exists()), None)
    if launcher is None:
        return False
    command = [str(launcher)]
    if json_path is not None:
        command.append(str(json_path))
    try:
        subprocess.Popen(
            command,
            cwd=str(launcher.parent),
            env=child_env(),
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            close_fds=True,
        )
        return True
    except Exception:
        return False


def main():
    app = QApplication(sys.argv)
    apply_light_palette(app)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
