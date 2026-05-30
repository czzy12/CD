import argparse
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from bankflow_v2.auto_detect import detect_bank_type
from bankflow_v2.generic_pdf import extract_generic_pdf
from bankflow_v2.pipeline import extract_transactions
from bankflow_v2.summary import money, summarize


PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")
IGNORE_FILL = PatternFill("solid", fgColor="E7E6E6")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

STATUS_LABELS = {
    "PASS": "通过",
    "GENERIC_PASS": "通用解析通过",
    "REVIEW": "需要复核",
    "GENERIC_REVIEW": "通用解析需复核",
    "UNRECOGNIZED": "未识别银行",
    "NO_ROWS": "未抽取到流水",
    "PARSE_ERROR": "解析失败",
    "IGNORE_IMAGE": "忽略-图片PDF",
    "IGNORE_PASSWORD": "忽略-加密PDF",
    "MANUAL_PASS": "已人工核对通过",
    "ERROR": "程序异常",
}

SUMMARY_HEADERS = [
    ("file", "文件"),
    ("status", "状态"),
    ("bank_id", "银行ID"),
    ("bank_label", "银行名称"),
    ("confidence", "识别置信度"),
    ("parser_id", "使用解析器"),
    ("count", "流水笔数"),
    ("income_count", "收入笔数"),
    ("income_sum", "收入合计"),
    ("expense_count", "支出笔数"),
    ("expense_sum", "支出合计"),
    ("net", "净额"),
    ("opening_balance", "期初余额"),
    ("closing_balance", "期末余额"),
    ("issue_count", "异常数"),
    ("first_issue", "首个异常"),
    ("detect_reason", "识别说明"),
    ("parse_error", "解析错误"),
]

ISSUE_HEADERS = [
    ("file", "文件"),
    ("status", "状态"),
    ("bank_id", "银行ID"),
    ("parser_id", "使用解析器"),
    ("time", "交易时间"),
    ("message", "异常说明"),
    ("raw_amount", "原始金额"),
    ("raw_balance", "原始余额"),
]


def _text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _status_fill(status: str) -> PatternFill:
    if "通过" in status:
        return PASS_FILL
    if status.startswith("忽略"):
        return IGNORE_FILL
    if "复核" in status or "未抽取" in status:
        return REVIEW_FILL
    return FAIL_FILL


def _status_label(status: str) -> str:
    return STATUS_LABELS.get(status, status)


def _is_image_pdf_reason(reason: str) -> bool:
    return any(keyword in reason for keyword in ("扫描", "图片", "OCR"))


def _try_parse(pdf_path: Path, bank_id: str):
    if bank_id:
        try:
            rows = extract_transactions(str(pdf_path), bank=bank_id)
            if rows:
                return rows, bank_id, ""
            generic_rows = extract_generic_pdf(str(pdf_path))
            if generic_rows:
                return generic_rows, "generic_pdf", f"{bank_id} returned 0 rows; generic fallback parsed rows"
            return rows, bank_id, ""
        except Exception as exc:
            generic_rows = []
            generic_error = ""
            try:
                generic_rows = extract_generic_pdf(str(pdf_path))
            except Exception as fallback_exc:
                generic_error = f"; generic fallback failed: {fallback_exc}"
            if generic_rows:
                return generic_rows, "generic_pdf", f"{bank_id} failed: {exc}; generic fallback parsed rows"
            return [], bank_id, f"{exc}{generic_error}"

    try:
        generic_rows = extract_generic_pdf(str(pdf_path))
    except Exception as exc:
        return [], "generic_pdf", f"generic fallback failed: {exc}"
    return generic_rows, "generic_pdf" if generic_rows else "", ""


def validate_one(pdf_path: Path, root: Path) -> tuple[dict, list[dict]]:
    relative_path = str(pdf_path.relative_to(root))
    detected = detect_bank_type(str(pdf_path))
    if not detected.bank_id and _is_image_pdf_reason(detected.reason):
        row = {
            "file": relative_path,
            "status": "IGNORE_IMAGE",
            "bank_id": "",
            "bank_label": detected.label,
            "confidence": detected.confidence,
            "parser_id": "",
            "count": 0,
            "income_count": 0,
            "income_sum": "",
            "expense_count": 0,
            "expense_sum": "",
            "net": "",
            "opening_balance": "",
            "closing_balance": "",
            "issue_count": 0,
            "first_issue": "",
            "detect_reason": detected.reason,
            "parse_error": "",
        }
        return row, []

    rows, parser_id, parse_error = _try_parse(pdf_path, detected.bank_id)
    summary = summarize(rows, source=pdf_path.name) if rows else None

    if parse_error and not rows:
        status = "PARSE_ERROR"
    elif not detected.bank_id and not rows:
        status = "UNRECOGNIZED"
    elif not rows:
        status = "NO_ROWS"
    elif summary and summary.issues:
        status = "GENERIC_REVIEW" if parser_id == "generic_pdf" else "REVIEW"
    else:
        status = "PASS" if parser_id != "generic_pdf" else "GENERIC_PASS"

    first_issue = ""
    issue_rows: list[dict] = []
    if summary:
        for issue in summary.issues:
            if not first_issue:
                first_issue = issue.message
            issue_rows.append(
                {
                    "file": relative_path,
                    "status": status,
                    "bank_id": detected.bank_id,
                    "parser_id": parser_id,
                    "time": issue.time,
                    "message": issue.message,
                    "raw_amount": issue.raw_amount,
                    "raw_balance": issue.raw_balance,
                }
            )

    row = {
        "file": relative_path,
        "status": status,
        "bank_id": detected.bank_id,
        "bank_label": detected.label,
        "confidence": detected.confidence,
        "parser_id": parser_id,
        "count": summary.count if summary else 0,
        "income_count": summary.income_count if summary else 0,
        "income_sum": money(summary.income_sum) if summary else "",
        "expense_count": summary.expense_count if summary else 0,
        "expense_sum": money(summary.expense_sum) if summary else "",
        "net": money(summary.net) if summary else "",
        "opening_balance": money(summary.opening_balance) if summary else "",
        "closing_balance": money(summary.closing_balance) if summary else "",
        "issue_count": len(summary.issues) if summary else 0,
        "first_issue": first_issue,
        "detect_reason": detected.reason,
        "parse_error": parse_error,
    }
    return row, issue_rows


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        max_len = max(len(_text(cell.value)) for cell in column_cells)
        width = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    status_col = None
    for index, cell in enumerate(ws[1]):
        if cell.value in {"status", "状态"}:
            status_col = index
            break
    for row in ws.iter_rows(min_row=2):
        fill = None
        if status_col is not None and len(row) > status_col:
            fill = _status_fill(_text(row[status_col].value))
        for cell in row:
            if fill is not None:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_report(output_path: Path, rows: list[dict], issues: list[dict], root: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    headers = [label for _, label in SUMMARY_HEADERS]
    ws.append(headers)
    for row in rows:
        values = []
        for key, _ in SUMMARY_HEADERS:
            value = row.get(key, "")
            if key == "status":
                value = _status_label(value)
            values.append(value)
        ws.append(values)
    _style_sheet(ws)

    issue_ws = wb.create_sheet("异常明细")
    issue_headers = [label for _, label in ISSUE_HEADERS]
    issue_ws.append(issue_headers)
    for issue in issues:
        values = []
        for key, _ in ISSUE_HEADERS:
            value = issue.get(key, "")
            if key == "status":
                value = _status_label(value)
            values.append(value)
        issue_ws.append(values)
    _style_sheet(issue_ws)

    counts = Counter(row["status"] for row in rows)
    overview_ws = wb.create_sheet("概览", 0)
    overview_ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview_ws.append(["扫描目录", str(root)])
    overview_ws.append(["PDF总数", len(rows)])
    overview_ws.append([])
    overview_ws.append(["状态", "数量"])
    for status, count in counts.most_common():
        overview_ws.append([_status_label(status), count])
    _style_sheet(overview_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Batch validate bank flow PDF files.")
    parser.add_argument("root", help="Directory containing PDF files.")
    parser.add_argument("--out", help="Output .xlsx path. Defaults to batch_validate.xlsx under root.")
    parser.add_argument(
        "--ignore-parse-errors",
        action="store_true",
        help="Mark parse errors as ignored encrypted/password PDFs.",
    )
    parser.add_argument(
        "--manual-pass",
        action="append",
        default=[],
        help="File name or relative path that has been manually checked and should be marked as pass.",
    )
    args = parser.parse_args()

    root = Path(args.root).expanduser().resolve()
    if not root.exists():
        raise SystemExit(f"Input directory does not exist: {root}")

    output_path = Path(args.out).expanduser().resolve() if args.out else root / "batch_validate.xlsx"
    pdf_files = sorted(root.rglob("*.pdf"))
    manual_pass = {item.casefold() for item in args.manual_pass}
    rows: list[dict] = []
    issues: list[dict] = []

    for index, pdf_path in enumerate(pdf_files, start=1):
        print(f"[{index}/{len(pdf_files)}] {pdf_path.name}")
        try:
            row, issue_rows = validate_one(pdf_path, root)
        except Exception as exc:
            row = {
                "file": str(pdf_path.relative_to(root)),
                "status": "ERROR",
                "bank_id": "",
                "bank_label": "",
                "confidence": 0,
                "parser_id": "",
                "count": 0,
                "income_count": 0,
                "income_sum": "",
                "expense_count": 0,
                "expense_sum": "",
                "net": "",
                "opening_balance": "",
                "closing_balance": "",
                "issue_count": 0,
                "first_issue": "",
                "detect_reason": "",
                "parse_error": str(exc),
            }
            issue_rows = []
        row_file = str(row["file"])
        if args.ignore_parse_errors and row["status"] == "PARSE_ERROR":
            row["status"] = "IGNORE_PASSWORD"
            row["first_issue"] = ""
            row["parse_error"] = row["parse_error"] or "需要密码，已从本轮测试剔除"
            issue_rows = []
        if row_file.casefold() in manual_pass or Path(row_file).name.casefold() in manual_pass:
            row["status"] = "MANUAL_PASS"
            row["first_issue"] = ""
            row["parse_error"] = ""
            issue_rows = []
        rows.append(row)
        issues.extend(issue_rows)

    write_report(output_path, rows, issues, root)
    print(f"Report: {output_path}")
    for status, count in Counter(row["status"] for row in rows).most_common():
        print(f"{_status_label(status)}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
