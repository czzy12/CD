import argparse
import re
import sys
from collections import Counter
from dataclasses import asdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from bankflow_v2 import extract_transactions


RAW_CMB_RE = re.compile(
    r"(?P<date>\d{4}-\d{2}-\d{2})\s+CNY\s+"
    r"(?P<amount>[+-]?\d[\d,]*\.\d{2})\s+"
    r"(?P<balance>\d[\d,]*\.\d{2})"
)
DATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")
MONEY_RE = re.compile(r"[+-]?\d[\d,]*\.\d{2}")


def money(value) -> Decimal:
    return Decimal(str(value).replace(",", "")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _money_candidates(value: Decimal | None) -> list[Decimal | None]:
    if value is None:
        return [None]
    candidates = [value]
    for divisor in (Decimal("10"), Decimal("100"), Decimal("1000"), Decimal("10000")):
        candidates.append((value / divisor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    return list(dict.fromkeys(candidates))


def repair_gold_continuity(rows: list[dict]) -> list[dict]:
    if not rows:
        return rows

    repaired = [rows[0]]
    for row in rows[1:]:
        prev_balance = repaired[-1].get("balance")
        if prev_balance is None or row.get("balance") is None:
            repaired.append(row)
            continue

        expected = (prev_balance + row["amount"]).quantize(Decimal("0.01"))
        if expected == row["balance"]:
            repaired.append(row)
            continue

        best_amount = row["amount"]
        best_balance = row["balance"]
        found = False
        for amount in _money_candidates(row["amount"]):
            if amount is None:
                continue
            for balance in _money_candidates(row["balance"]):
                if balance is None:
                    continue
                if (prev_balance + amount).quantize(Decimal("0.01")) == balance:
                    best_amount = amount
                    best_balance = balance
                    found = True
                    break
            if found:
                break

        if found:
            row = dict(row)
            row["amount"] = best_amount
            row["income"] = best_amount if best_amount > 0 else Decimal("0.00")
            row["expense"] = -best_amount if best_amount < 0 else Decimal("0.00")
            row["balance"] = best_balance
        repaired.append(row)

    return repaired


def load_gold(excel_path: str, sheet_name: str = "Sheet1") -> list[dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    rows: list[dict] = []

    header_row = 1
    headers = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True), start=1):
        labels = [str(value or "").strip().replace("\n", "") for value in row]
        if (
            ("交易日期" in labels or "交易时间" in labels or "记账日期" in labels)
            and (
                "收入/支出金额" in labels
                or "交易金额" in labels
                or "借" in labels
                or "贷" in labels
            )
        ):
            header_row = idx
            headers = labels
            break

    if not headers:
        headers = [str(cell.value or "").strip().replace("\n", "") for cell in ws[1]]

    date_idx = next(
        (headers.index(name) for name in ("交易日期", "交易时间", "记账日期") if name in headers),
        0,
    )
    time_idx = headers.index("交易时间") if "交易时间" in headers else None
    amount_idx = next(
        (headers.index(name) for name in ("收入/支出金额", "交易金额") if name in headers),
        None,
    )
    debit_idx = headers.index("借") if "借" in headers else None
    credit_idx = headers.index("贷") if "贷" in headers else None
    balance_idx = next(
        (headers.index(name) for name in ("余额", "交易余额", "联机余额", "本次余额") if name in headers),
        None,
    )

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= date_idx or not row[date_idx]:
            continue

        if not hasattr(row[date_idx], "date"):
            raw_text = str(row[date_idx])
            if re.fullmatch(r"\d{8}", raw_text) and amount_idx is not None:
                raw_time = (
                    str(row[time_idx]).zfill(6)
                    if time_idx is not None and len(row) > time_idx and row[time_idx] not in (None, "")
                    else "000000"
                )
                if not re.fullmatch(r"\d{6}", raw_time):
                    raw_time = "000000"
                amount_value = row[amount_idx] if len(row) > amount_idx else None
                balance_value = row[balance_idx] if balance_idx is not None and len(row) > balance_idx else None
                if amount_value not in (None, ""):
                    tx_time = datetime(
                        int(raw_text[:4]),
                        int(raw_text[4:6]),
                        int(raw_text[6:8]),
                        int(raw_time[:2]),
                        int(raw_time[2:4]),
                        int(raw_time[4:6]),
                    )
                    amount = money(amount_value)
                    balance = money(balance_value) if balance_value not in (None, "") else None
                    rows.append(
                        {
                            "transaction_time": tx_time,
                            "amount": amount,
                            "income": amount if amount > 0 else Decimal("0.00"),
                            "expense": -amount if amount < 0 else Decimal("0.00"),
                            "balance": balance,
                        }
                    )
                    continue

            dates = DATE_RE.findall(raw_text)
            amount_values = MONEY_RE.findall(str(row[amount_idx - 1] if amount_idx and len(row) > amount_idx - 1 else ""))
            balance_values = MONEY_RE.findall(str(row[balance_idx] if balance_idx is not None and len(row) > balance_idx else ""))
            if len(dates) > 1 and len(dates) == len(amount_values) == len(balance_values):
                for raw_date, raw_amount, raw_balance in zip(dates, amount_values, balance_values):
                    tx_time = datetime.strptime(raw_date, "%Y-%m-%d")
                    amount = money(raw_amount)
                    balance = money(raw_balance)
                    rows.append(
                        {
                            "transaction_time": tx_time,
                            "amount": amount,
                            "income": amount if amount > 0 else Decimal("0.00"),
                            "expense": -amount if amount < 0 else Decimal("0.00"),
                            "balance": balance,
                        }
                    )
                continue

            if len(dates) == 1 and amount_idx is not None and balance_idx is not None:
                amount_value = row[amount_idx] if len(row) > amount_idx else None
                if amount_value in (None, "") and amount_idx > 0 and len(row) > amount_idx - 1:
                    amount_value = row[amount_idx - 1]
                balance_value = row[balance_idx] if len(row) > balance_idx else None
                if amount_value not in (None, "") and balance_value not in (None, ""):
                    tx_time = datetime.strptime(dates[0], "%Y-%m-%d")
                    amount = money(amount_value)
                    balance = money(balance_value)
                    rows.append(
                        {
                            "transaction_time": tx_time,
                            "amount": amount,
                            "income": amount if amount > 0 else Decimal("0.00"),
                            "expense": -amount if amount < 0 else Decimal("0.00"),
                            "balance": balance,
                        }
                    )
                    continue

            match = RAW_CMB_RE.search(raw_text)
            if not match:
                continue
            tx_time = datetime.strptime(match.group("date"), "%Y-%m-%d")
            amount = money(match.group("amount"))
            balance = money(match.group("balance"))
            rows.append(
                {
                    "transaction_time": tx_time,
                    "amount": amount,
                    "income": amount if amount > 0 else Decimal("0.00"),
                    "expense": -amount if amount < 0 else Decimal("0.00"),
                    "balance": balance,
                }
            )
            continue

        if amount_idx is not None:
            amount_value = row[amount_idx] if len(row) > amount_idx else None
            if amount_value in (None, "") and amount_idx > 0 and len(row) > amount_idx - 1:
                amount_value = row[amount_idx - 1]
            if amount_value in (None, ""):
                continue
            amount = money(amount_value)
        elif debit_idx is not None and credit_idx is not None:
            debit = money(row[debit_idx] or 0)
            credit = money(row[credit_idx] or 0)
            if debit == 0 and credit == 0:
                continue
            amount = credit - debit
        else:
            continue

        balance = (
            money(row[balance_idx])
            if balance_idx is not None and len(row) > balance_idx and row[balance_idx] is not None
            else None
        )
        rows.append(
            {
                "transaction_time": row[date_idx],
                "amount": amount,
                "income": amount if amount > 0 else Decimal("0.00"),
                "expense": -amount if amount < 0 else Decimal("0.00"),
                "balance": balance,
            }
        )
    return repair_gold_continuity(rows)


def filter_range(transactions, start: datetime, end: datetime):
    first = min(start, end)
    last = max(start, end)
    start_day = datetime(first.year, first.month, first.day)
    end_day = datetime(last.year, last.month, last.day, 23, 59, 59, 999999)
    return [tx for tx in transactions if start_day <= tx.transaction_time <= end_day]


def summarize(rows: list[dict]) -> dict:
    income_count = sum(
        1
        for row in rows
        if row.get("amount") is not None and row["amount"] >= 0
    )
    expense_count = sum(1 for row in rows if row["expense"] > 0)
    income_sum = sum((row["income"] for row in rows), Decimal("0.00"))
    expense_sum = sum((row["expense"] for row in rows), Decimal("0.00"))
    return {
        "count": len(rows),
        "income_count": income_count,
        "income_sum": income_sum,
        "expense_count": expense_count,
        "expense_sum": expense_sum,
        "net": income_sum - expense_sum,
        "first_balance": rows[0].get("balance") if rows else None,
        "last_balance": rows[-1].get("balance") if rows else None,
    }


def tx_to_row(tx) -> dict:
    return {
        "transaction_time": tx.transaction_time,
        "amount": tx.amount,
        "income": tx.income,
        "expense": tx.expense,
        "balance": tx.balance,
        "status": tx.status,
        "issues": tx.issues,
        "raw_time": tx.raw_time,
        "raw_amount": tx.raw_amount,
        "raw_balance": tx.raw_balance,
    }


def compare(gold_rows: list[dict], parsed_rows: list[dict]) -> list[str]:
    diffs: list[str] = []
    if len(gold_rows) != len(parsed_rows):
        diffs.append(f"笔数不一致: gold={len(gold_rows)} parsed={len(parsed_rows)}")

    for idx, (gold, parsed) in enumerate(zip(gold_rows, parsed_rows), start=1):
        checks = ["amount", "income", "expense"]
        if gold.get("balance") is not None:
            checks.append("balance")
        for key in checks:
            if gold[key] != parsed[key]:
                diffs.append(
                    f"第 {idx} 行 {key} 不一致: gold={gold[key]} parsed={parsed[key]} "
                    f"raw=({parsed.get('raw_time')} | {parsed.get('raw_amount')} | {parsed.get('raw_balance')})"
                )
        if gold["transaction_time"].date() != parsed["transaction_time"].date():
            diffs.append(
                f"第 {idx} 行日期不一致: gold={gold['transaction_time']} parsed={parsed['transaction_time']}"
            )
    return diffs


def main():
    parser = argparse.ArgumentParser(description="黄金样本对账")
    parser.add_argument("--pdf", default=r"C:\Users\czzy1\Desktop\流水测试\流水测试\陈洁银行卡.pdf")
    parser.add_argument("--gold", default=r"C:\Users\czzy1\Desktop\陈洁银行卡.xlsx")
    parser.add_argument("--bank", default="icbc")
    args = parser.parse_args()

    gold_rows = load_gold(args.gold)
    start = gold_rows[0]["transaction_time"]
    end = gold_rows[-1]["transaction_time"]

    parsed = extract_transactions(args.pdf, bank=args.bank)
    parsed_rows = [tx_to_row(tx) for tx in filter_range(parsed, start, end)]

    print("PDF:", Path(args.pdf))
    print("Gold:", Path(args.gold))
    print("gold summary:", summarize(gold_rows))
    print("parsed summary:", summarize(parsed_rows))
    print("statuses:", Counter(row["status"] for row in parsed_rows))

    diffs = compare(gold_rows, parsed_rows)
    if diffs:
        print("\nDIFFS")
        for diff in diffs[:50]:
            print("-", diff)
        raise SystemExit(1)

    review_rows = [row for row in parsed_rows if row["status"] != "ok"]
    if review_rows:
        print("\nREVIEW")
        for row in review_rows[:20]:
            print(asdict(row) if hasattr(row, "__dataclass_fields__") else row)

    print("\nPASS")


if __name__ == "__main__":
    main()
