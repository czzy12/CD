"""Gate F3B: Human Gold review prep, validation and freeze (no prediction).

AI never fills gold. This module only prepares review files, validates
human-filled enums/refs/invariants, and freezes the final Human Gold.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import random
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from bankflow_v2.knowledge.freeze import manifest_checksum


REVIEW_STANDARD_VERSION = "human_gold_review_standard_v1"
RELATION_VALUES = ("strong", "medium", "weak", "none", "undetermined")
ROLE_VALUES = (
    "direct_business",
    "operating_expense",
    "tax_regulatory",
    "financing",
    "settlement_infrastructure",
    "employment_operation",
    "government_interaction",
    "personal_consumption",
    "neutral_transfer",
    "unknown",
)
TRACE_VALUES = ("strong", "medium", "weak", "none", "undetermined")
ROUTE_VALUES = (
    "local_resolved",
    "ai_eligible_transaction",
    "insufficient_transaction",
    "case_aggregation_only",
)
PRESENCE_VALUES = ("strong", "medium", "weak", "none", "undetermined")
CONSISTENCY_VALUES = ("strong", "medium", "weak", "none", "undetermined")
SUFFICIENCY_VALUES = ("sufficient", "partial", "insufficient")
CONFIDENCE_VALUES = ("high", "medium", "low")

TX_GOLD_COLUMNS = (
    "human_industry_direct_relation",
    "human_business_evidence_role",
    "human_business_trace_strength",
    "human_expected_route",
    "human_sufficient_information",
    "human_confidence",
    "supporting_evidence_refs",
    "reviewer_reasoning",
    "reviewer_id",
    "reviewed_at",
    "review_standard_version",
)

CASE_GOLD_COLUMNS = (
    "business_activity_presence",
    "declared_industry_consistency",
    "human_assessment_sufficiency",
    "supporting_evidence_refs",
    "contradictory_evidence_refs",
    "uncertainty_notes",
    "reasoning_summary",
    "reviewer_id",
    "reviewed_at",
    "review_standard_version",
)

TX_REVIEW_INFO_COLUMNS = (
    "holdout_item_id",
    "anonymized_case_id",
    "declared_industry",
    "business_description",
    "normalized_transaction_text",
    "safe_semantic_evidence",
    "date",
    "month",
    "direction",
    "amount",
    "amount_bucket",
    "evidence_refs",
)

TX_HEADER_ZH = {
    "holdout_item_id": "交易编号",
    "anonymized_case_id": "案例ID(匿名)",
    "declared_industry": "申报行业",
    "business_description": "业务描述",
    "normalized_transaction_text": "交易归一化文本",
    "safe_semantic_evidence": "安全语义证据(JSON)",
    "date": "交易日期",
    "month": "月份",
    "direction": "方向",
    "amount": "金额",
    "amount_bucket": "金额档",
    "evidence_refs": "证据引用",
    "human_industry_direct_relation": "行业直接关系(人工)",
    "human_business_evidence_role": "经营证据角色(人工)",
    "human_business_trace_strength": "经营痕迹强度(人工)",
    "human_expected_route": "预期处理层(人工)",
    "human_sufficient_information": "信息是否充分(人工)",
    "human_confidence": "人工置信度",
    "supporting_evidence_refs": "支持证据引用(人工)",
    "reviewer_reasoning": "判断理由(人工)",
    "reviewer_id": "审核人",
    "reviewed_at": "审核时间",
    "review_standard_version": "审核标准版本",
}

CASE_HEADER_ZH = {
    "anonymized_case_id": "案例ID(匿名)",
    "declared_industry": "申报行业",
    "business_description": "业务描述",
    "account_source_coverage": "账户/来源覆盖",
    "company_address_available": "公司地址可用",
    "home_address_available": "家庭地址可用",
    "business_activity_presence": "经营存在(人工)",
    "declared_industry_consistency": "申报行业一致性(人工)",
    "human_assessment_sufficiency": "资料充分性(人工)",
    "supporting_evidence_refs": "支持证据引用(人工)",
    "contradictory_evidence_refs": "矛盾证据引用(人工)",
    "uncertainty_notes": "不确定说明(人工)",
    "reasoning_summary": "判断逻辑(人工)",
    "reviewer_id": "审核人",
    "reviewed_at": "审核时间",
    "review_standard_version": "审核标准版本",
}

QC_HEADER_ZH = {
    "qc_item_id": "复核编号",
    "holdout_item_id": "原交易编号",
    "declared_industry": "申报行业",
    "normalized_transaction_text": "交易归一化文本",
    "safe_semantic_evidence": "安全语义证据(JSON)",
    "date": "交易日期",
    "direction": "方向",
    "amount": "金额",
    "blank_for_rereview": "复核用空白",
}

RELATION_ZH = {
    "strong": "强",
    "medium": "中",
    "weak": "弱",
    "none": "无",
    "undetermined": "无法判断",
}
ROLE_ZH = {
    "direct_business": "直接经营",
    "operating_expense": "经营运营支出",
    "tax_regulatory": "税务监管",
    "financing": "融资借贷",
    "settlement_infrastructure": "结算基础设施",
    "employment_operation": "用工经营",
    "government_interaction": "政府往来",
    "personal_consumption": "个人消费",
    "neutral_transfer": "中性转账",
    "unknown": "未知",
}
ROUTE_ZH = {
    "local_resolved": "本地处理",
    "ai_eligible_transaction": "AI处理",
    "insufficient_transaction": "信息不足",
    "case_aggregation_only": "仅案件汇总",
}
SUFFICIENCY_ZH = {
    "sufficient": "充分",
    "partial": "部分",
    "insufficient": "不足",
}
CONFIDENCE_ZH = {
    "high": "高",
    "medium": "中",
    "low": "低",
}
BOOL_ZH = {"true": "是", "false": "否"}

ZH_TO_EN = {}
for _map in (RELATION_ZH, ROLE_ZH, ROUTE_ZH, SUFFICIENCY_ZH, CONFIDENCE_ZH, BOOL_ZH):
    ZH_TO_EN.update({zh: en for en, zh in _map.items()})


def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sha256_file(path: Path) -> str:
    return _sha256_bytes(path.read_bytes())


def load_holdout_items(holdout_dir: Path) -> list[dict[str, Any]]:
    path = holdout_dir / "production_transaction_evidence_holdout_v1_items.jsonl"
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def load_case_meta(holdout_dir: Path, output_dir: Path) -> list[dict[str, Any]]:
    case_manifest = json.loads(
        (holdout_dir / "production_case_holdout_v1_manifest.json").read_text(
            encoding="utf-8"
        )
    )
    inventory = json.loads(
        (holdout_dir / "0808_case_inventory.json").read_text(encoding="utf-8")
    )
    by_id = {row["anonymized_case_id"]: row for row in inventory}
    rows: list[dict[str, Any]] = []
    for case_id in case_manifest["case_ids"]:
        row = by_id.get(case_id, {})
        rows.append(
            {
                "anonymized_case_id": case_id,
                "declared_industry": row.get("declared_industry", ""),
                "business_description": row.get("business_description", ""),
                "source_directory": row.get("source_directory", ""),
                "statement_files": row.get("statement_files", []),
                "company_address_available": row.get(
                    "company_address_available",
                    False,
                ),
                "home_address_available": row.get(
                    "home_address_available",
                    False,
                ),
            }
        )
    return rows


def build_transaction_review_csv(items: list[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=TX_REVIEW_INFO_COLUMNS + TX_GOLD_COLUMNS,
        extrasaction="ignore",
    )
    writer.writeheader()
    for item in sorted(items, key=lambda row: row["holdout_item_id"]):
        writer.writerow(
            {
                "holdout_item_id": item["holdout_item_id"],
                "anonymized_case_id": item["source_case_id"],
                "declared_industry": item["declared_industry"],
                "business_description": item["declared_industry"],
                "normalized_transaction_text": item[
                    "normalized_transaction_text"
                ],
                "safe_semantic_evidence": json.dumps(
                    item["safe_semantic_evidence"],
                    ensure_ascii=False,
                ),
                "date": item["date"],
                "month": item["month"],
                "direction": item["direction"],
                "amount": item["amount"],
                "amount_bucket": item["amount_bucket"],
                "evidence_refs": item["source_evidence_reference"],
                **{column: "" for column in TX_GOLD_COLUMNS},
            }
        )
    return buffer.getvalue()


def build_case_review_csv(cases: list[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=(
            "anonymized_case_id",
            "declared_industry",
            "business_description",
            "account_source_coverage",
            "company_address_available",
            "home_address_available",
        )
        + CASE_GOLD_COLUMNS,
        extrasaction="ignore",
    )
    writer.writeheader()
    for case in sorted(cases, key=lambda row: row["anonymized_case_id"]):
        writer.writerow(
            {
                "anonymized_case_id": case["anonymized_case_id"],
                "declared_industry": case["declared_industry"],
                "business_description": case["business_description"],
                "account_source_coverage": "|".join(
                    str(doc.get("bank_id", ""))
                    for doc in case["statement_files"]
                    if doc.get("supported")
                ),
                "company_address_available": case[
                    "company_address_available"
                ],
                "home_address_available": case["home_address_available"],
                **{column: "" for column in CASE_GOLD_COLUMNS},
            }
        )
    return buffer.getvalue()


def build_qc_list(items: list[dict[str, Any]], *, count: int = 10) -> str:
    rng = random.Random(20260808)
    selected = sorted(items, key=lambda row: row["holdout_item_id"])
    rng.shuffle(selected)
    selected = selected[:count]
    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=(
            "qc_item_id",
            "holdout_item_id",
            "declared_industry",
            "normalized_transaction_text",
            "safe_semantic_evidence",
            "date",
            "direction",
            "amount",
            "blank_for_rereview",
        ),
    )
    writer.writeheader()
    for index, item in enumerate(selected, start=1):
        writer.writerow(
            {
                "qc_item_id": f"QC-{index:02d}",
                "holdout_item_id": item["holdout_item_id"],
                "declared_industry": item["declared_industry"],
                "normalized_transaction_text": item[
                    "normalized_transaction_text"
                ],
                "safe_semantic_evidence": json.dumps(
                    item["safe_semantic_evidence"],
                    ensure_ascii=False,
                ),
                "date": item["date"],
                "direction": item["direction"],
                "amount": item["amount"],
                "blank_for_rereview": "true",
            }
        )
    return buffer.getvalue()


def write_chinese_header_copy(
    source: Path,
    target: Path,
    header_map: dict[str, str],
) -> None:
    with open(source, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    translated = [header_map.get(name, name) for name in fieldnames]
    with open(target, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=translated)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    header_map.get(name, name): row[name]
                    for name in fieldnames
                }
            )


def _add_dropdown(ws, column_letter: str, options: list[str], row_count: int) -> None:
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{column_letter}2:{column_letter}{row_count + 1}")


def build_transaction_xlsx(items: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "交易人工审核"
    headers = [TX_HEADER_ZH[name] for name in TX_REVIEW_INFO_COLUMNS + TX_GOLD_COLUMNS]
    ws.append(headers)
    for item in sorted(items, key=lambda row: row["holdout_item_id"]):
        ws.append(
            [
                item["holdout_item_id"],
                item["source_case_id"],
                item["declared_industry"],
                item["declared_industry"],
                item["normalized_transaction_text"],
                json.dumps(item["safe_semantic_evidence"], ensure_ascii=False),
                item["date"],
                item["month"],
                item["direction"],
                item["amount"],
                item["amount_bucket"],
                item["source_evidence_reference"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "human_gold_review_standard_v1",
            ]
        )
    row_count = len(items)
    _add_dropdown(ws, "M", list(RELATION_ZH.values()), row_count)
    _add_dropdown(ws, "N", list(ROLE_ZH.values()), row_count)
    _add_dropdown(ws, "O", list(RELATION_ZH.values()), row_count)
    _add_dropdown(ws, "P", list(ROUTE_ZH.values()), row_count)
    _add_dropdown(ws, "Q", ["是", "否"], row_count)
    _add_dropdown(ws, "R", list(CONFIDENCE_ZH.values()), row_count)
    _add_dropdown(ws, "W", ["human_gold_review_standard_v1"], row_count)
    wb.save(path)


def build_case_xlsx(cases: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "案件人工审核"
    info_headers = (
        "anonymized_case_id",
        "declared_industry",
        "business_description",
        "account_source_coverage",
        "company_address_available",
        "home_address_available",
    )
    headers = [CASE_HEADER_ZH[name] for name in info_headers + CASE_GOLD_COLUMNS]
    ws.append(headers)
    for case in sorted(cases, key=lambda row: row["anonymized_case_id"]):
        ws.append(
            [
                case["anonymized_case_id"],
                case["declared_industry"],
                case["business_description"],
                "|".join(
                    str(doc.get("bank_id", ""))
                    for doc in case["statement_files"]
                    if doc.get("supported")
                ),
                "是" if case["company_address_available"] else "否",
                "是" if case["home_address_available"] else "否",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "human_gold_review_standard_v1",
            ]
        )
    row_count = len(cases)
    _add_dropdown(ws, "G", list(RELATION_ZH.values()), row_count)
    _add_dropdown(ws, "H", list(RELATION_ZH.values()), row_count)
    _add_dropdown(ws, "I", list(SUFFICIENCY_ZH.values()), row_count)
    _add_dropdown(ws, "P", ["human_gold_review_standard_v1"], row_count)
    wb.save(path)


def build_qc_xlsx(items: list[dict[str, Any]], path: Path, count: int = 10) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "QC复核"
    rng = random.Random(20260808)
    selected = sorted(items, key=lambda row: row["holdout_item_id"])
    rng.shuffle(selected)
    selected = selected[:count]
    headers = [
        "复核编号",
        "原交易编号",
        "申报行业",
        "交易归一化文本",
        "安全语义证据(JSON)",
        "交易日期",
        "方向",
        "金额",
        "行业直接关系(人工)",
        "经营证据角色(人工)",
        "经营痕迹强度(人工)",
        "预期处理层(人工)",
        "信息是否充分(人工)",
        "人工置信度",
        "判断理由(人工)",
        "审核人",
        "审核时间",
        "审核标准版本",
    ]
    ws.append(headers)
    for index, item in enumerate(selected, start=1):
        ws.append(
            [
                f"QC-{index:02d}",
                item["holdout_item_id"],
                item["declared_industry"],
                item["normalized_transaction_text"],
                json.dumps(item["safe_semantic_evidence"], ensure_ascii=False),
                item["date"],
                item["direction"],
                item["amount"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "human_gold_review_standard_v1",
            ]
        )
    _add_dropdown(ws, "I", list(RELATION_ZH.values()), count)
    _add_dropdown(ws, "J", list(ROLE_ZH.values()), count)
    _add_dropdown(ws, "K", list(RELATION_ZH.values()), count)
    _add_dropdown(ws, "L", list(ROUTE_ZH.values()), count)
    _add_dropdown(ws, "M", ["是", "否"], count)
    _add_dropdown(ws, "N", list(CONFIDENCE_ZH.values()), count)
    _add_dropdown(ws, "R", ["human_gold_review_standard_v1"], count)
    wb.save(path)


def read_transaction_gold_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "") for cell in ws[1]]
    reverse_headers = {zh: en for en, zh in TX_HEADER_ZH.items()}
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        record: dict[str, Any] = {}
        for header, value in zip(headers, row):
            english = reverse_headers.get(header, header)
            if english in TX_GOLD_COLUMNS or english == "holdout_item_id":
                record[english] = ZH_TO_EN.get(str(value or ""), str(value or ""))
        rows.append(record)
    return rows


def read_case_gold_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "") for cell in ws[1]]
    reverse_headers = {zh: en for en, zh in CASE_HEADER_ZH.items()}
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        record: dict[str, Any] = {}
        for header, value in zip(headers, row):
            english = reverse_headers.get(header, header)
            if english in CASE_GOLD_COLUMNS or english == "anonymized_case_id":
                record[english] = ZH_TO_EN.get(str(value or ""), str(value or ""))
        rows.append(record)
    return rows


def validate_transaction_gold(
    gold: list[dict[str, Any]],
    items: list[dict[str, Any]],
) -> dict[str, Any]:
    by_id = {item["holdout_item_id"]: item for item in items}
    errors: list[str] = []
    warnings: list[str] = []
    for row in gold:
        item_id = str(row.get("holdout_item_id", ""))
        if item_id not in by_id:
            errors.append(f"{item_id}: unknown holdout item")
            continue
        if str(row.get("human_industry_direct_relation", "")) not in RELATION_VALUES:
            errors.append(f"{item_id}: invalid industry relation")
        if str(row.get("human_business_evidence_role", "")) not in ROLE_VALUES:
            errors.append(f"{item_id}: invalid evidence role")
        if str(row.get("human_business_trace_strength", "")) not in TRACE_VALUES:
            errors.append(f"{item_id}: invalid trace strength")
        if str(row.get("human_expected_route", "")) not in ROUTE_VALUES:
            errors.append(f"{item_id}: invalid route")
        if not str(row.get("reviewer_reasoning", "")).strip():
            errors.append(f"{item_id}: reviewer_reasoning missing")
        if not str(row.get("reviewer_id", "")).strip():
            errors.append(f"{item_id}: reviewer_id missing")
        if (
            row.get("human_industry_direct_relation") == "none"
            and row.get("human_sufficient_information") != "true"
        ):
            warnings.append(
                f"{item_id}: none used without sufficient information"
            )
        if (
            row.get("human_business_evidence_role") == "unknown"
            and row.get("human_business_trace_strength")
            not in {"undetermined", ""}
        ):
            warnings.append(
                f"{item_id}: unknown role with non-undetermined trace"
            )
    return {
        "reviewed_count": len(gold),
        "expected_count": len(items),
        "complete": len(gold) == len(items),
        "errors": errors,
        "warnings": warnings,
    }


def validate_case_gold(
    case_gold: list[dict[str, Any]],
    cases: list[dict[str, Any]],
) -> dict[str, Any]:
    case_ids = {case["anonymized_case_id"] for case in cases}
    errors: list[str] = []
    warnings: list[str] = []
    for row in case_gold:
        case_id = str(row.get("anonymized_case_id", ""))
        if case_id not in case_ids:
            errors.append(f"{case_id}: unknown case")
            continue
        if str(row.get("business_activity_presence", "")) not in PRESENCE_VALUES:
            errors.append(f"{case_id}: invalid presence")
        if (
            str(row.get("declared_industry_consistency", ""))
            not in CONSISTENCY_VALUES
        ):
            errors.append(f"{case_id}: invalid consistency")
        if str(row.get("human_assessment_sufficiency", "")) not in SUFFICIENCY_VALUES:
            errors.append(f"{case_id}: invalid sufficiency")
        if not str(row.get("reasoning_summary", "")).strip():
            errors.append(f"{case_id}: reasoning_summary missing")
        if not str(row.get("reviewer_id", "")).strip():
            errors.append(f"{case_id}: reviewer_id missing")
        if (
            row.get("declared_industry_consistency") == "none"
            and row.get("business_activity_presence") in {"none", "undetermined"}
        ):
            warnings.append(
                f"{case_id}: consistency none with weak/no activity presence"
            )
    return {
        "reviewed_count": len(case_gold),
        "expected_count": len(cases),
        "complete": len(case_gold) == len(cases),
        "errors": errors,
        "warnings": warnings,
    }


def freeze_gold(
    *,
    output_dir: Path,
    tx_gold: list[dict[str, Any]],
    case_gold: list[dict[str, Any]],
    qc_results: dict[str, Any],
    reviewer: str,
    reviewed_at_range: list[str],
    candidate_v2_checksum: str,
    holdout_checksum: str,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    tx_path = output_dir / "production_transaction_evidence_human_gold_v1.jsonl"
    case_path = output_dir / "production_case_human_gold_v1.json"
    tx_path.write_text(
        "".join(
            json.dumps(row, ensure_ascii=False) + "\n" for row in tx_gold
        ),
        encoding="utf-8",
    )
    case_path.write_text(
        json.dumps(case_gold, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    manifest = {
        "freeze_version": "human-gold-v1",
        "review_standard_version": REVIEW_STANDARD_VERSION,
        "reviewer": reviewer,
        "reviewed_at_range": reviewed_at_range,
        "transaction_item_count": len(tx_gold),
        "case_count": len(case_gold),
        "artifact_checksums": {
            "transaction_gold": _sha256_file(tx_path),
            "case_gold": _sha256_file(case_path),
        },
        "qc_results": qc_results,
        "candidate_v2_checksum": candidate_v2_checksum,
        "holdout_checksum": holdout_checksum,
        "prediction_call_count": 0,
        "provider_call_count": 0,
        "created_at": _utcnow(),
    }
    manifest["aggregate_checksum"] = manifest_checksum(manifest)
    manifest_path = output_dir / "human_gold_freeze_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--holdout-dir",
        type=Path,
        default=Path(
            "D:/Investigator PDF/outputs/knowledge-v1/"
            "gate-f3a-1-resume-holdout-20260808"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(
            "D:/Investigator PDF/outputs/knowledge-v1/"
            "gate-f3b-human-gold-20260809"
        ),
    )
    parser.add_argument("--qc-count", type=int, default=10)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    items = load_holdout_items(args.holdout_dir)
    cases = load_case_meta(args.holdout_dir, args.output_dir)
    (args.output_dir / "transaction_human_review_v1.csv").write_text(
        build_transaction_review_csv(items),
        encoding="utf-8-sig",
    )
    (args.output_dir / "case_human_review_v1.csv").write_text(
        build_case_review_csv(cases),
        encoding="utf-8-sig",
    )
    (args.output_dir / "transaction_qc_rereview_v1.csv").write_text(
        build_qc_list(items, count=args.qc_count),
        encoding="utf-8-sig",
    )
    (args.output_dir / "case_review_meta.json").write_text(
        json.dumps(cases, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_chinese_header_copy(
        args.output_dir / "transaction_human_review_v1.csv",
        args.output_dir / "transaction_human_review_v1_中文版.csv",
        TX_HEADER_ZH,
    )
    write_chinese_header_copy(
        args.output_dir / "case_human_review_v1.csv",
        args.output_dir / "case_human_review_v1_中文版.csv",
        CASE_HEADER_ZH,
    )
    write_chinese_header_copy(
        args.output_dir / "transaction_qc_rereview_v1.csv",
        args.output_dir / "transaction_qc_rereview_v1_中文版.csv",
        QC_HEADER_ZH,
    )
    build_transaction_xlsx(
        items,
        args.output_dir / "transaction_human_review_v1_下拉填写版.xlsx",
    )
    build_case_xlsx(
        cases,
        args.output_dir / "case_human_review_v1_下拉填写版.xlsx",
    )
    build_qc_xlsx(
        items,
        args.output_dir / "transaction_qc_rereview_v1_下拉填写版.xlsx",
        count=args.qc_count,
    )
    report = {
        "gate": "F3B-PREP",
        "review_standard_version": REVIEW_STANDARD_VERSION,
        "transaction_review_rows": len(items),
        "case_review_rows": len(cases),
        "qc_count": args.qc_count,
        "gold_status": "blank",
        "prediction_call_count": 0,
        "provider_call_count": 0,
        "created_at": _utcnow(),
    }
    (args.output_dir / "f3b_prep_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print("status=prepared")
    print(f"transaction_review_rows={len(items)}")
    print(f"case_review_rows={len(cases)}")
    print(f"qc_count={args.qc_count}")
    print(f"output={args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
